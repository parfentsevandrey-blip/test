"""Разбор xlsx-реестра в нормализованный registry.json.

Реестр вёлся для людей, а не для машины: этаж записан строкой «6 из 11», отделка и
машино-место спрятаны в текст комментария. Скрипт вытаскивает это регулярками и честно
помечает, что выведено эвристикой, а что взято из явной колонки, — список inferred полей
показывает, каких полей реестру не хватает для оценки.

    python scripts/parse_registry.py data/reestr.xlsx data/registry.json
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from agent.models import Finish  # noqa: E402

HEADER_ROW_MARKER = "ЖК"

# Порядок важен: первое совпадение выигрывает, поэтому deluxe стоит выше дизайнерского.
FINISH_PATTERNS: list[tuple[str, Finish]] = [
    (r"deluxe|натурального камня", Finish.DELUXE),
    (r"дизайнерск|авторск\w+ интерьер", Finish.DESIGNER),
    (r"отделк\w+ от застройщика", Finish.DEVELOPER),
    (r"индивидуальн\w+ дизайн-проект|white ?box|под отделку", Finish.WHITEBOX),
]

# «машино-место можно приобрести за N млн» — место продаётся отдельно, в лот не входит.
PARKING_SEPARATE = re.compile(r"машино-мест\w*\s+можно\s+приобрести", re.I)
PARKING_INCLUDED = re.compile(r"машино-мест", re.I)

FLOOR_RE = re.compile(r"(\d+)\s*из\s*(\d+)")
LANDING_RE = re.compile(r"https?://[^\s,]+")


def parse_floor(raw: str) -> tuple[int, int]:
    m = FLOOR_RE.search(str(raw))
    if not m:
        raise ValueError(f"не разобран этаж: {raw!r}")
    return int(m.group(1)), int(m.group(2))


def infer_finish(comment: str) -> tuple[Finish, bool]:
    """→ (отделка, выведено_ли_эвристикой). Без совпадений считаем «от застройщика»."""
    for pattern, finish in FINISH_PATTERNS:
        if re.search(pattern, comment, re.I):
            return finish, True
    return Finish.DEVELOPER, True


def infer_parking(comment: str) -> tuple[bool, bool]:
    if PARKING_SEPARATE.search(comment):
        return False, True
    if PARKING_INCLUDED.search(comment):
        return True, True
    return False, True


def slugify(complex_name: str, index: int) -> str:
    translit = str.maketrans(
        "абвгдеёжзийклмнопрстуфхцчшщъыьэюя",
        "abvgdeejziyklmnoprstufhccss_y_eua",
    )
    base = complex_name.strip().lower().translate(translit)
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    return f"{base or 'lot'}-{index}"


def parse(xlsx_path: Path) -> list[dict]:
    from openpyxl import load_workbook

    ws = load_workbook(xlsx_path, data_only=True).worksheets[0]

    header_row = next(
        (
            row[0].row
            for row in ws.iter_rows()
            if str(row[0].value).strip() == HEADER_ROW_MARKER
        ),
        None,
    )
    if header_row is None:
        raise ValueError("не найдена строка заголовков (ожидалась ячейка «ЖК»)")

    lots: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        complex_name, address, rooms, area, floor_raw, comment, video, deck, price = row[:9]
        # Реестр заканчивается строкой-подвалом без цены — она же и признак конца таблицы.
        if not complex_name or not isinstance(price, (int, float)):
            continue

        comment = (comment or "").strip()
        floor, floors_total = parse_floor(floor_raw)
        finish, finish_inferred = infer_finish(comment)
        has_parking, parking_inferred = infer_parking(comment)
        landing = LANDING_RE.search(comment)

        inferred = ["finish"] * finish_inferred + ["has_parking"] * parking_inferred
        lots.append(
            {
                "id": slugify(str(complex_name), len(lots) + 1),
                "complex_name": str(complex_name).strip(),
                "address": str(address).strip(),
                "rooms": int(rooms),
                "area": float(area),
                "floor": floor,
                "floors_total": floors_total,
                "price": int(price),
                "finish": finish.value,
                "has_parking": has_parking,
                "comment": comment,
                "video_url": (str(video).strip() if video and str(video).strip() != "-" else ""),
                "presentation_url": (str(deck).strip() if deck else ""),
                "landing_url": landing.group(0).rstrip(".,") if landing else "",
                # Полей ниже в реестре нет — их должна отдавать CRM/фид площадки.
                "listed_at": None,
                "views_7d": None,
                "calls_7d": None,
                "viewings_30d": None,
                "last_price_change": None,
                "price_history": [],
                "_inferred_fields": inferred,
            }
        )
    return lots


def main() -> None:
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "data/reestr.xlsx")
    dst = Path(sys.argv[2] if len(sys.argv) > 2 else "data/registry.json")

    lots = parse(src)
    dst.write_text(json.dumps(lots, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Разобрано лотов: {len(lots)} → {dst}")
    for lot in lots:
        print(
            f"  {lot['id']:<28} {lot['price'] / lot['area'] / 1000:>8.0f} тыс ₽/м²  "
            f"{lot['finish']:<26} машино-место: {'да' if lot['has_parking'] else 'нет'}"
        )
    print(
        "\nВыведено эвристикой из текста комментария (нужно подтвердить вручную): "
        "отделка, машино-место."
    )
    print(
        "Отсутствует в реестре и требуется для вердикта: дата выхода в экспозицию, "
        "история цен, просмотры/звонки/показы."
    )


if __name__ == "__main__":
    main()
