"""Яндекс Недвижимость: три канала, каждый со своей ролью.

У площадки нет партнёрского REST API, через который можно было бы вытянуть чужие
объявления или статистику. Зато есть три работающих канала, и вместе они закрывают
главный пробел реестра — данные о реакции рынка на нашу цену:

  1. YRL-фид (исходящий).  Мы сами публикуем объявления XML-фидом, который площадка
     забирает ежедневно. Здесь же назначается `internal-id` — и именно он потом
     приезжает обратно в выгрузке звонков. Это и есть канал записи новой цены.

  2. Выгрузка из кабинета юрлица (входящий).  Кнопка «скачать» отдаёт XLS с логом
     звонков: дата и время, подменный номер, добавочный, длительность, стоимость
     звонка, тип сделки, тип объекта, ID из XML-фида, регион, статус. Одна строка —
     один звонок, поэтому спрос по лоту считается агрегацией по `internal-id`.
     Автоматической выгрузки нет: файл кладут в папку руками, по расписанию из CRM
     или роботом в кабинете.

  3. Оценка квартиры (контрольная точка).  ML-калькулятор площадки по адресу,
     комнатности, этажу, площади и состоянию ремонта. Публичного API у него нет,
     поэтому адаптер ниже — каркас: путь метода задаётся переменной окружения и
     подставляется из договорённости с площадкой. Пока не сконфигурирован —
     возвращает None, и ядро просто считает коридор без сверки.

Чего Яндекс НЕ даёт (как и Циан): выгрузки чужих объявлений. Аналоги для коридора
берутся из ЕГРН, своей базы и платной аналитики — см. docs/IMPLEMENTATION.md.
"""

from __future__ import annotations

import csv
import logging
import os
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterator

import httpx

from ..models import Apartment, Finish, HouseValuation, OpsSnapshot

log = logging.getLogger(__name__)

# --- 1. Выгрузка звонков из кабинета ----------------------------------------------

# Заголовки в выгрузке могут отличаться от версии к версии кабинета, поэтому
# колонки ищутся по подстроке, а не по точному совпадению.
COLUMN_HINTS: dict[str, tuple[str, ...]] = {
    "feed_id": ("id из xml", "xml", "идентификатор объявления", "внутренний id"),
    "at": ("дата", "время"),
    "duration": ("длительность", "продолжительность"),
    "status": ("статус",),
}

# Пропущенные и нулевые по длительности звонки спросом не считаем.
MIN_CALL_SECONDS = 5
FAILED_STATUSES = ("пропущ", "не отвеч", "недозвон", "сброш")


class YandexCallLog:
    """Спрос по лотам из XLS/CSV-выгрузки звонков кабинета юрлица."""

    name = "yandex-cabinet"

    def __init__(self, path: Path | str | None = None) -> None:
        raw = path or os.getenv("YANDEX_CALLS_EXPORT", "")
        self.path = Path(raw) if raw else None
        self._by_feed_id: dict[str, list[datetime]] | None = None

    @property
    def available(self) -> bool:
        return self.path is not None and self.path.exists()

    def fetch_ops(self, apartment: Apartment) -> OpsSnapshot | None:
        """Звонки за 7 и 30 дней по конкретному лоту."""
        if not self.available:
            return None
        calls = self._index().get(apartment.feed_id)
        if not calls:
            # Объявление есть в реестре, но звонков в выгрузке нет — это тоже сигнал,
            # а не отсутствие данных: значит за период не позвонили ни разу.
            return OpsSnapshot(source=self.name, calls_7d=0, calls_30d=0)

        now = datetime.now()
        return OpsSnapshot(
            source=self.name,
            calls_7d=sum(1 for t in calls if now - t <= timedelta(days=7)),
            calls_30d=sum(1 for t in calls if now - t <= timedelta(days=30)),
        )

    def _index(self) -> dict[str, list[datetime]]:
        if self._by_feed_id is None:
            self._by_feed_id = self._build_index()
        return self._by_feed_id

    def _build_index(self) -> dict[str, list[datetime]]:
        try:
            rows = list(self._rows())
        except Exception as exc:  # битый файл не должен ронять бота
            log.error("Не разобрана выгрузка звонков %s: %s", self.path, exc)
            return {}

        if not rows:
            return {}

        header = [str(c or "").strip().lower() for c in rows[0]]
        cols = {key: _find_column(header, hints) for key, hints in COLUMN_HINTS.items()}
        if cols["feed_id"] is None or cols["at"] is None:
            log.error(
                "В выгрузке %s нет колонок с ID объявления и датой — "
                "проверьте, что скачан лог звонков, а не сводка",
                self.path,
            )
            return {}

        index: dict[str, list[datetime]] = defaultdict(list)
        skipped = 0
        for row in rows[1:]:
            feed_id = _cell(row, cols["feed_id"])
            when = _parse_dt(_cell(row, cols["at"]))
            if not feed_id or when is None:
                skipped += 1
                continue
            if _is_failed_call(_cell(row, cols["status"]), _cell(row, cols["duration"])):
                continue
            index[feed_id].append(when)

        log.info(
            "Выгрузка звонков: %d объявлений, %d строк пропущено",
            len(index),
            skipped,
        )
        return dict(index)

    def _rows(self) -> Iterator[list[Any]]:
        assert self.path is not None
        if self.path.suffix.lower() in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            ws = load_workbook(self.path, data_only=True, read_only=True).worksheets[0]
            yield from ws.iter_rows(values_only=True)
            return

        # Кабинет отдаёт XLS, но выгрузку часто пересохраняют в CSV — поддерживаем оба.
        with self.path.open(encoding="utf-8-sig", newline="") as fh:
            header = fh.readline()
            fh.seek(0)
            yield from csv.reader(fh, delimiter=_sniff_delimiter(header))


def _sniff_delimiter(header_line: str) -> str:
    """Разделитель определяется по строке заголовков, а не по всему файлу.

    csv.Sniffer здесь ошибается: в русских выгрузках деньги пишутся с десятичной
    запятой («450,00»), и по телу файла запятых оказывается больше, чем настоящих
    разделителей. Строка заголовков цифр не содержит, поэтому по ней счёт честный.
    """
    return max(";,\t", key=header_line.count)


def _find_column(header: list[str], hints: tuple[str, ...]) -> int | None:
    for i, title in enumerate(header):
        if any(h in title for h in hints):
            return i
    return None


def _cell(row: list[Any], idx: int | None) -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _is_failed_call(status: str, duration: str) -> bool:
    if any(bad in status.lower() for bad in FAILED_STATUSES):
        return True
    seconds = re.sub(r"[^\d]", "", duration)
    return bool(seconds) and int(seconds) < MIN_CALL_SECONDS


DT_FORMATS = (
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d",
)
# Длина строки, которую даёт каждый формат — по ней обрезается «хвост» вроде часового
# пояса или миллисекунд, если кабинет их добавит.
_DT_WIDTHS = {f: len(datetime(2026, 12, 31, 23, 59, 59).strftime(f)) for f in DT_FORMATS}


def _parse_dt(value: str) -> datetime | None:
    if not value:
        return None
    value = value.replace("T", " ").strip()
    for fmt in DT_FORMATS:
        for candidate in (value, value[: _DT_WIDTHS[fmt]]):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


# --- 2. Оценка квартиры как контрольная точка -------------------------------------

VALUATION_URL = os.getenv("YANDEX_VALUATION_URL", "")


class YandexValuation:
    """Независимая оценка ₽/м² по дому.

    Каркас, а не готовая интеграция: публичного API у калькулятора нет, конкретный
    адрес метода и схема ответа приходят по договорённости с площадкой. Форма
    интеграции здесь зафиксирована — что отправляем, что маппим, как деградируем.
    """

    name = "yandex-valuation"

    def __init__(self, timeout: float = 10.0) -> None:
        self.timeout = timeout
        self._warned = False

    @property
    def configured(self) -> bool:
        return bool(VALUATION_URL)

    def fetch_valuation(self, apartment: Apartment) -> HouseValuation | None:
        if not self.configured:
            if not self._warned:
                log.warning(
                    "Оценка Яндекса не сконфигурирована (YANDEX_VALUATION_URL) — "
                    "коридор считается без сверки"
                )
                self._warned = True
            return None

        payload = {
            "address": apartment.address,
            "rooms": apartment.rooms,
            "area": apartment.area,
            "floor": apartment.floor,
            "floors_total": apartment.floors_total,
            "renovation": RENOVATION[apartment.finish],
        }
        try:
            resp = httpx.post(VALUATION_URL, json=payload, timeout=self.timeout)
            resp.raise_for_status()
            data = resp.json()
        except (httpx.HTTPError, ValueError) as exc:
            log.error("Оценка по дому не получена: %s", exc)
            return None

        ppsm = _extract_price_per_sqm(data, apartment.area)
        if ppsm is None:
            log.error("В ответе оценки нет цены — проверьте маппинг схемы")
            return None

        return HouseValuation(
            source=self.name,
            price_per_sqm=ppsm,
            observed_at=date.today(),
            note="ML-оценка площадки по дому",
        )


def _extract_price_per_sqm(data: dict[str, Any], area: float) -> float | None:
    """Ответ может отдавать как цену за метр, так и цену лота — принимаем оба."""
    for key in ("price_per_sqm", "pricePerSqm", "square_meter_price"):
        if isinstance(data.get(key), (int, float)):
            return float(data[key])
    for key in ("price", "value", "estimate"):
        if isinstance(data.get(key), (int, float)) and area > 0:
            return float(data[key]) / area
    return None


# --- 3. Генерация YRL-фида --------------------------------------------------------

YRL_NS = "http://webmaster.yandex.ru/schemas/feed/realty/2010-06"

RENOVATION: dict[Finish, str] = {
    Finish.NONE: "черновая",
    Finish.WHITEBOX: "требует ремонта",
    Finish.DEVELOPER: "чистовая",
    Finish.DESIGNER: "дизайнерский ремонт",
    Finish.DELUXE: "дизайнерский ремонт",
}


def build_yrl_feed(
    apartments: list[Apartment],
    *,
    prices: dict[str, int] | None = None,
    organization: str = "",
    phone: str = "",
    generated_at: datetime | None = None,
) -> str:
    """Собирает YRL-фид для выгрузки на площадку.

    `prices` — переопределение цен по id лота: сюда кладут утверждённые рекомендации,
    и фид становится каналом применения новой цены. Без него берётся цена из реестра.

    `internal-id` каждого объявления равен `apartment.feed_id` — по нему выгрузка
    звонков из кабинета связывается обратно с объектом реестра.

    Перед первой боевой выгрузкой фид нужно прогнать через валидатор площадки:
    набор обязательных элементов у YRL со временем меняется.
    """
    import xml.etree.ElementTree as ET

    prices = prices or {}
    root = ET.Element("realty-feed", {"xmlns": YRL_NS})
    stamp = (generated_at or datetime.now()).replace(microsecond=0)
    ET.SubElement(root, "generation-date").text = stamp.isoformat()

    for a in apartments:
        offer = ET.SubElement(root, "offer", {"internal-id": a.feed_id})
        ET.SubElement(offer, "type").text = "продажа"
        ET.SubElement(offer, "property-type").text = "жилая"
        ET.SubElement(offer, "category").text = "квартира"

        location = ET.SubElement(offer, "location")
        ET.SubElement(location, "country").text = "Россия"
        ET.SubElement(location, "region").text = "Москва"
        ET.SubElement(location, "address").text = a.address

        if organization or phone:
            agent = ET.SubElement(offer, "sales-agent")
            if phone:
                ET.SubElement(agent, "phone").text = phone
            ET.SubElement(agent, "category").text = "агентство"
            if organization:
                ET.SubElement(agent, "organization").text = organization

        price = ET.SubElement(offer, "price")
        ET.SubElement(price, "value").text = str(prices.get(a.id, a.price))
        ET.SubElement(price, "currency").text = "RUR"

        area = ET.SubElement(offer, "area")
        ET.SubElement(area, "value").text = f"{a.area:g}"
        ET.SubElement(area, "unit").text = "кв. м"

        ET.SubElement(offer, "rooms").text = str(a.rooms)
        ET.SubElement(offer, "floor").text = str(a.floor)
        ET.SubElement(offer, "floors-total").text = str(a.floors_total)
        ET.SubElement(offer, "building-name").text = a.complex_name
        ET.SubElement(offer, "renovation").text = RENOVATION[a.finish]
        if a.comment:
            ET.SubElement(offer, "description").text = a.comment

    ET.indent(root, space="  ")
    return '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(root, encoding="unicode")
