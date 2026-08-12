"""Импорт выгрузок расширения «Циан → Excel» как источника аналогов.

Это единственный источник в проекте, который даёт **чужие лоты** — то, чего не отдают
ни партнёрское API Циан, ни Яндекс Недвижимость. Расширение собирает выдачу по ЖК или
по фильтру и складывает её в .xlsx; здесь этот файл превращается в аналоги для коридора.

Что из выгрузки реально ценно для ядра, помимо цены и площади:

  «Реальный срок, дн»   — срок экспозиции с поправкой на переподачи. Счётчик Циан
                          сбрасывается при переподаче, и лот, который висит полгода,
                          выглядит свежим. Наш вердикт опирается на срок экспозиции,
                          поэтому это самая важная колонка после цены.
  «Δ цены с 1-й выгрузки» — накопленное снижение цены конкурентом. Ровно тот сигнал,
                          по которому ядро понимает, что рынок торгуется вниз.
  «Переподач»           — сколько раз объявление перевыставляли. Много переподач при
                          длинной экспозиции = лот стоит и продавец это скрывает.
  «Отделка/ремонт»      — с указанием, определена она полем Циан или вытащена из текста.

Формат листа «Все_лоты» задаётся расширением: строки 1–3 — заголовок, подзаголовок и
примечания, строка 4 — названия колонок, данные с 5-й. Колонки ищутся по названию, а не
по индексу: расширение развивается, порядок и состав колонок меняются между версиями.
"""

from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from ..location import canonical_key, load_projects, normalise
from ..models import Apartment, Comp, Finish

log = logging.getLogger(__name__)

SHEET = "Все_лоты"
HEADER_ROW = 4

# Словарь отделки расширения → наша порядковая шкала.
# «Под ключ / с мебелью» и «Евроремонт» приравнены к дизайнерскому: по вкладу в цену
# они ближе к нему, чем к чистовой отделке от застройщика.
FINISH_MAP: dict[str, Finish] = {
    "без отделки": Finish.NONE,
    "черновая": Finish.NONE,
    "без ремонта": Finish.NONE,
    "предчистовая (white box)": Finish.WHITEBOX,
    "чистовая": Finish.DEVELOPER,
    "косметический": Finish.DEVELOPER,
    "с ремонтом (тип не указан)": Finish.DEVELOPER,
    "евроремонт": Finish.DESIGNER,
    "под ключ / с мебелью": Finish.DESIGNER,
    "дизайнерский": Finish.DESIGNER,
}

COLUMNS = {
    "area": ("площадь",),
    "floor": ("этаж",),
    "building": ("корпус",),
    "seller_type": ("тип продавца",),
    "seller_name": ("продавец",),
    "finish": ("отделка",),
    "price": ("цена, ₽",),
    "ppsm": ("цена за м",),
    "delta_first": ("δ цены с 1-й", "с 1-й выгрузки"),
    "exposure": ("реальный срок",),
    "republish": ("переподач",),
    "url": ("ссылка",),
}

FLOOR_RE = re.compile(r"(\d+)\s*/\s*(\d+)")
DEVELOPER = "застройщик"


class CianExportProvider:
    """Аналоги из выгрузок расширения. Одна выгрузка — один ЖК."""

    name = "cian-export"

    def __init__(
        self,
        directory: Path | str,
        *,
        include_developer: bool = False,
        max_age_days: int = 45,
    ) -> None:
        self.directory = Path(directory)
        self.include_developer = include_developer
        self.max_age_days = max_age_days
        self._by_complex: dict[str, list[Comp]] | None = None
        self._own: list[Apartment] = []
        self._projects = load_projects()
        self.stats: dict[str, Any] = {}

    def exclude_own(self, apartments: list[Apartment]) -> None:
        """Запоминает весь наш портфель, а не один оцениваемый лот.

        В Sky House у нас три квартиры, и все три висят на Циан. Без этого списка
        каждая попадала бы в выборку двух других: объект сравнивался бы сам с собой,
        коридор тянулся бы к нашему же прайсу, и завышение целой группы лотов стало бы
        для агента невидимым — он подтвердил бы его как «рынок».
        """
        self._own = list(apartments)

    @property
    def available(self) -> bool:
        return self.directory.is_dir() and any(self._files())

    def _files(self) -> Iterable[Path]:
        if not self.directory.is_dir():
            return []
        # rglob, а не glob: выгрузки удобно раскладывать по подпапкам-локациям.
        return sorted(p for p in self.directory.rglob("*.xlsx") if not p.name.startswith("~$"))

    def house_lots(self, apartment: Apartment) -> list[Comp]:
        """Вся экспозиция нашего ЖК, включая прайс застройщика. Без наших лотов.

        Отдельно от fetch_comps, потому что у прайса застройщика есть применение, для
        которого он незаменим: по нему меряется надбавка за этаж. Это единственный
        набор цен в доме, где все прочие условия равны по построению — один продавец,
        одна отделка, один момент времени, — поэтому регрессия на нём даёт чистый
        коэффициент, а на разрозненных частных объявлениях разваливается.
        """
        index = self._index()
        comps = index.get(self._key(apartment.complex_name), [])
        if not comps:
            comps = _fuzzy_lookup(index, apartment.complex_name)
        own = self._own or [apartment]
        return [c for c in comps if not any(_is_same_lot(c, o) for o in own)]

    def fetch_comps(self, apartment: Apartment, radius_km: float = 1.5) -> list[Comp]:
        """Аналоги для коридора: свой ЖК, без лотов застройщика.

        Первичка и вторичка — разные рынки, и для нашего лота перепродажи прайс
        застройщика смещает коридор. Для АНАЛИЗА ЛОКАЦИИ фильтр обратный: там
        новостройка и есть конкуренция, поэтому by_project() отдаёт всё —
        иначе из выборки исчезает целый конкурирующий проект.
        """
        return [
            c
            for c in self.house_lots(apartment)
            if self.include_developer or DEVELOPER not in (c.seller_type or "").lower()
        ]

    def by_project(self) -> dict[str, list[Comp]]:
        """Все выгрузки, сгруппированные по ЖК — сырьё для анализа локации."""
        return dict(self._index())

    def _key(self, name: str) -> str:
        """Ключ проекта с учётом алиасов: «Level» и «Левел» — один и тот же ЖК."""
        return canonical_key(name, self._projects)

    def _index(self) -> dict[str, list[Comp]]:
        if self._by_complex is None:
            self._by_complex = self._build_index()
        return self._by_complex

    def _build_index(self) -> dict[str, list[Comp]]:
        index: dict[str, list[Comp]] = {}
        totals = {"files": 0, "rows": 0, "kept": 0, "developer": 0, "stale": 0, "bad": 0}

        for path in self._files():
            try:
                complex_name, comps, counters = self._read(path)
            except Exception as exc:
                log.error("Выгрузка %s не разобрана: %s", path.name, exc)
                continue

            totals["files"] += 1
            for key, value in counters.items():
                totals[key] += value
            index.setdefault(self._key(complex_name), []).extend(comps)
            log.info(
                "%s: ЖК «%s», принято %d аналогов из %d строк",
                path.name,
                complex_name,
                len(comps),
                counters["rows"],
            )

        # Один и тот же лот встречается и в разных выгрузках, и внутри одной —
        # расширение считает такие дубли, но не схлопывает их.
        for key, comps in index.items():
            index[key] = _dedupe(comps)

        self.stats = totals
        return index

    def _read(self, path: Path) -> tuple[str, list[Comp], dict[str, int]]:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        if SHEET not in wb.sheetnames:
            raise ValueError(f"нет листа «{SHEET}» — это не выгрузка расширения")
        ws = wb[SHEET]

        # A1 расширение пишет как «<название ЖК> — все лоты».
        title = str(ws.cell(row=1, column=1).value or "")
        complex_name = title.split("—")[0].strip() or path.stem

        header = [
            str(ws.cell(row=HEADER_ROW, column=i).value or "").strip().lower()
            for i in range(1, ws.max_column + 1)
        ]
        cols = {key: _find(header, hints) for key, hints in COLUMNS.items()}
        missing = [k for k in ("area", "floor", "price") if cols[k] is None]
        if missing:
            raise ValueError(f"в листе нет обязательных колонок: {', '.join(missing)}")

        observed = date.fromtimestamp(path.stat().st_mtime)
        comps: list[Comp] = []
        counters = {"rows": 0, "kept": 0, "developer": 0, "stale": 0, "bad": 0}

        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            cell = lambda key: ws.cell(row=r, column=cols[key]) if cols[key] else None  # noqa: E731
            area = _number(_value(cell("area")))
            price = _number(_value(cell("price")))
            if area is None and price is None:
                continue  # хвост листа
            counters["rows"] += 1

            floor, floors_total = _parse_floor(_value(cell("floor")))
            if not area or not price or floor is None:
                counters["bad"] += 1
                continue

            seller = str(_value(cell("seller_type")) or "").strip()
            if DEVELOPER in seller.lower():
                counters["developer"] += 1

            exposure = _number(_value(cell("exposure")))
            comps.append(
                Comp(
                    source=self.name,
                    external_id=f"{path.stem}:{r}",
                    complex_name=complex_name,
                    address=str(_value(cell("building")) or complex_name).strip(),
                    rooms=0,  # комнатность в выгрузке текстовая («2-комн»), ядру не нужна
                    area=area,
                    floor=floor,
                    floors_total=floors_total or floor,
                    price=int(price),
                    finish=_parse_finish(_value(cell("finish"))),
                    has_parking=False,  # расширение машино-место не выделяет
                    same_complex=True,  # выгрузка делается по конкретному ЖК
                    distance_km=0.0,
                    days_on_market=int(exposure) if exposure else None,
                    price_cut_pct=_parse_delta(_value(cell("delta_first"))),
                    observed_at=observed,
                    url=_hyperlink(cell("url")),
                    seller_type=seller,
                    seller_name=str(_value(cell("seller_name")) or "").strip(),
                    republish=_int(_value(cell("republish"))),
                )
            )

        fresh = [c for c in comps if (date.today() - observed).days <= self.max_age_days]
        counters["stale"] = len(comps) - len(fresh)
        counters["kept"] = len(fresh)
        if counters["stale"]:
            log.warning(
                "%s: выгрузка от %s старше %d дней — %d аналогов пропущено",
                path.name,
                observed,
                self.max_age_days,
                counters["stale"],
            )
        return complex_name, fresh, counters

    def summary(self) -> str:
        # Индекс строится лениво, а сводку спрашивают до первого расчёта — без этого
        # вызова она отчиталась бы, что выгрузок нет, хотя они уже используются.
        self._index()
        s = self.stats
        if not s.get("files"):
            return "Выгрузок расширения не найдено."
        parts = [f"выгрузок {s['files']}", f"строк {s['rows']}", f"аналогов {s['kept']}"]
        if s["developer"]:
            parts.append(
                f"лотов застройщика {s['developer']} "
                f"({'учтены' if self.include_developer else 'вне коридора, но в локации'})"
            )
        if s["stale"]:
            parts.append(f"устаревших {s['stale']}")
        if s["bad"]:
            parts.append(f"нечитаемых строк {s['bad']}")
        return "Расширение «Циан → Excel»: " + ", ".join(parts)


# --- сопоставление ЖК --------------------------------------------------------------


def _fuzzy_lookup(index: dict[str, list[Comp]], complex_name: str) -> list[Comp]:
    """Запасное сопоставление: «Золотой жилой квартал» ↔ «Золотой квартал».

    Берётся только однозначное совпадение: если под условие подходят два ЖК, лучше
    отдать пусто и показать «аналогов нет», чем молча смешать два разных дома.
    """
    target = normalise(complex_name)
    hits = [
        comps
        for key, comps in index.items()
        if key and (key in target or target in key)
    ]
    return hits[0] if len(hits) == 1 else []


# --- разбор ячеек ------------------------------------------------------------------


def _find(header: list[str], hints: tuple[str, ...]) -> int | None:
    """Точное совпадение имеет приоритет над подстрокой.

    Иначе «Продавец» матчится на «Тип продавца» — заголовок стоит левее и содержит
    искомую подстроку, и имя конкурента молча подменяется словом «агентство».
    """
    for i, title in enumerate(header, start=1):
        if any(title == h for h in hints):
            return i
    for i, title in enumerate(header, start=1):
        if any(h in title for h in hints):
            return i
    return None


def _value(cell: Any) -> Any:
    return None if cell is None else cell.value


def _hyperlink(cell: Any) -> str:
    """Ссылка лежит в гиперссылке ячейки, а не в её тексте («Циан →»)."""
    if cell is None:
        return ""
    if getattr(cell, "hyperlink", None) is not None:
        return str(cell.hyperlink.target or "")
    value = str(cell.value or "")
    return value if value.startswith("http") else ""


def _number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if not value:
        return None
    # Excel-выгрузка может отдать «12 500 000» с неразрывными пробелами.
    cleaned = re.sub(r"[^\d,.\-]", "", str(value)).replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _int(value: Any) -> int | None:
    n = _number(value)
    return int(n) if n is not None else None


def _parse_floor(value: Any) -> tuple[int | None, int | None]:
    """Расширение пишет этаж как «9/29», иногда «9» или «?/29»."""
    text = str(value or "").strip()
    m = FLOOR_RE.search(text)
    if m:
        return int(m.group(1)), int(m.group(2))
    n = _int(text)
    return (n, None) if n else (None, None)


def _parse_finish(value: Any) -> Finish:
    key = str(value or "").strip().lower().replace("ё", "е")
    for known, finish in FINISH_MAP.items():
        if key == known.replace("ё", "е"):
            return finish
    # Отделка не определена — берём середину шкалы, иначе поправка систематически
    # утянет коридор в одну сторону.
    return Finish.DEVELOPER


def _parse_delta(value: Any) -> float | None:
    """«Δ цены с 1-й выгрузки» приходит строкой вида «−6%» или «+3%».

    Ядру нужно накопленное СНИЖЕНИЕ как положительная доля, поэтому рост цены
    возвращается как None, а не как отрицательное снижение.
    """
    if value is None or value == "":
        return None
    text = str(value).replace("−", "-").replace("—", "-").strip()
    n = _number(text)
    if n is None or n >= 0:
        return None
    return abs(n) / 100


def _is_same_lot(comp: Comp, apartment: Apartment) -> bool:
    """Совпадение по ЖК, этажу и площади. Цена намеренно не проверяется.

    Выгрузка могла быть сделана до нашего последнего изменения цены, и тогда сверка
    по цене не сработала бы. Ошибиться здесь безопаснее в сторону исключения: потерять
    один аналог — мелочь, сравнить лот сам с собой — систематическая ошибка.
    """
    return (
        normalise(comp.complex_name) == normalise(apartment.complex_name)
        and comp.floor == apartment.floor
        and abs(comp.area - apartment.area) < 0.6
    )


def _dedupe(comps: list[Comp]) -> list[Comp]:
    """Один лот, выставленный несколькими агентствами, не должен голосовать дважды."""
    seen: dict[tuple, Comp] = {}
    for c in comps:
        key = (c.floor, round(c.area * 2) / 2, round(c.price / 500_000))
        current = seen.get(key)
        # Из дублей оставляем тот, у кого длиннее реальный срок экспозиции: он ближе
        # к настоящей дате выхода лота на рынок.
        if current is None or (c.days_on_market or 0) > (current.days_on_market or 0):
            seen[key] = c
    return list(seen.values())
