#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
cian_scraper.py — выгрузка ВСЕХ активных квартир выбранного ЖК с cian.ru в Excel.

Идея: фронтенд Циан тянет выдачу POST-запросом на недокументированный JSON-API
    https://api.cian.ru/search-offers/v2/search-offers-desktop/
с телом {"jsonQuery": {...}}. Фильтр по ЖК — ключ "newobject".

ВАЖНО про схему запроса
-----------------------
Структура тела (jsonQuery) и набор заголовков у Циан со временем меняются, а доступ
часто требует ваших браузерных cookie. Поэтому скрипт умеет ДВА режима:

  1) Встроенный шаблон запроса (DEFAULT_JSON_QUERY ниже) — отражает устоявшуюся
     схему search-offers-desktop. Подходит, если API отвечает без капчи.

  2) --curl-file PATH — «Copy as cURL» реального запроса из DevTools
     (Network → XHR → search-offers). Скрипт разберёт ваш cURL и возьмёт ОТТУДА
     заголовки, cookie и тело jsonQuery как есть, подставляя только newobject / page /
     room / price. Это самый надёжный путь и именно он рекомендуется.

Колонки Excel, антибан, пагинация с обходом лимита 28×28, дедуп по cianId,
fallback на Playwright — см. README_cian.md и --help.
"""

import argparse
import json
import logging
import os
import random
import re
import shlex
import sys
import time
import unicodedata
from datetime import datetime, date, timezone

import requests

# ----------------------------------------------------------------------------- #
#  КОНСТАНТЫ (всё переопределяемо через CLI / --curl-file)                       #
# ----------------------------------------------------------------------------- #

API_URL = "https://api.cian.ru/search-offers/v2/search-offers-desktop/"

# Десктопный набор заголовков «как у браузера». Referer выставляется на лету.
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
    "Accept-Language": "ru,en;q=0.9",
    "Content-Type": "application/json",
    "Origin": "https://www.cian.ru",
}

# Коды комнатности в API Циан (ключ "room").  Подтвердите по своему cURL.
#   1..6 — число комнат (6 = «6 и более»), 7 — свободная планировка, 9 — студия.
ROOM_CODES = {
    "studio": 9,
    "openplan": 7,
    "1": 1,
    "2": 2,
    "3": 3,
    "4": 4,
    "5": 5,
    "6": 6,
}
# Порядок обхода категорий при room-split (умножает «потолок» 784 на число категорий).
ROOM_SPLIT_ORDER = [9, 7, 1, 2, 3, 4, 5, 6]

# Человекочитаемая категория для Excel по (room_code / признакам объявления).
CAT_STUDIO = "Студия"
CAT_OPENPLAN = "Своб. планировка"

PAGE_SIZE = 28          # Циан отдаёт ~28 лотов на страницу
PUBLIC_PAGE_CAP = 28    # и режет выдачу ~28 страницами (≈784 лота на запрос)

DUPLICATE_FLAG_IN = 1   # значение колонки «В расчёте» (1 — учитывать в средних)

log = logging.getLogger("cian")


# ----------------------------------------------------------------------------- #
#  РАЗБОР ВХОДА: ID ЖК из любой ссылки/значения                                  #
# ----------------------------------------------------------------------------- #

def extract_newobject_id(value):
    """
    Достаёт ID ЖК (newobject) из:
      • https://www.cian.ru/zhiloy-kompleks-symphony-34-2515016/  -> 2515016
      • ...?newobject%5B0%5D=2515016 / ?newobject[0]=2515016 / &newobject=2515016
      • ...-2515016/ (любой хвостовой числовой ID в slug)
      • просто "2515016"
    Возвращает int или бросает ValueError.
    """
    if value is None:
        raise ValueError("Не задан ЖК (--jk).")
    s = str(value).strip()

    if s.isdigit():
        return int(s)

    # query-параметр newobject[...]=ID  (в т.ч. URL-энкод %5B0%5D)
    m = re.search(r"newobject(?:%5[bB]\d+%5[dD]|\[\d*\])?=(\d+)", s)
    if m:
        return int(m.group(1))

    # хвостовой ID в slug пути:  -2515016/  или  -2515016?
    m = re.search(r"-(\d+)(?:/|\?|$)", s)
    if m:
        return int(m.group(1))

    # любой длинный (>=5 цифр) числовой кусок как последний шанс
    nums = re.findall(r"\d{5,}", s)
    if nums:
        return int(nums[-1])

    raise ValueError(f"Не удалось извлечь ID ЖК из: {value!r}")


# ----------------------------------------------------------------------------- #
#  РАЗБОР «Copy as cURL» из DevTools                                             #
# ----------------------------------------------------------------------------- #

def parse_curl(curl_text):
    """
    Разбирает строку cURL (bash-форма из Chrome/Firefox DevTools) и возвращает
    dict: {"url", "headers": {...}, "cookies": {...}, "json": {...}|None}.

    Поддерживает -H/--header, -b/--cookie, --data/--data-raw/--data-binary/-d.
    Переносы строк с обратным слэшем склеиваются.
    """
    text = curl_text.strip()
    # склейка переносов "\<newline>"
    text = re.sub(r"\\\s*\n", " ", text)
    # уберём ведущее слово curl, если есть
    text = re.sub(r"^\s*curl\s+", "", text, count=1)

    try:
        tokens = shlex.split(text)
    except ValueError:
        # на случай экзотического экранирования — грубый фолбэк
        tokens = text.split()

    url = None
    headers = {}
    cookies = {}
    data = None

    i = 0
    while i < len(tokens):
        t = tokens[i]
        if t in ("-H", "--header") and i + 1 < len(tokens):
            h = tokens[i + 1]
            if ":" in h:
                k, v = h.split(":", 1)
                headers[k.strip()] = v.strip()
            i += 2
        elif t in ("-b", "--cookie") and i + 1 < len(tokens):
            cookies.update(_parse_cookie_string(tokens[i + 1]))
            i += 2
        elif t in ("--data", "--data-raw", "--data-binary", "--data-ascii", "-d") and i + 1 < len(tokens):
            data = tokens[i + 1]
            i += 2
        elif t in ("--compressed", "-s", "-S", "-k", "--insecure", "-L", "--location", "-i", "-v"):
            i += 1
        elif t in ("-X", "--request", "-A", "--user-agent", "-e", "--referer") and i + 1 < len(tokens):
            # -A/-e продублируем как заголовки
            if t in ("-A", "--user-agent"):
                headers["User-Agent"] = tokens[i + 1]
            elif t in ("-e", "--referer"):
                headers["Referer"] = tokens[i + 1]
            i += 2
        elif t.startswith("http://") or t.startswith("https://"):
            url = t
            i += 1
        else:
            i += 1

    # cookie мог приехать заголовком Cookie:
    for hk in list(headers.keys()):
        if hk.lower() == "cookie":
            cookies.update(_parse_cookie_string(headers.pop(hk)))

    parsed_json = None
    if data:
        try:
            parsed_json = json.loads(data)
        except json.JSONDecodeError:
            log.warning("Тело cURL не распарсилось как JSON — будет использован шаблон по умолчанию.")
    return {"url": url, "headers": headers, "cookies": cookies, "json": parsed_json}


def _parse_cookie_string(s):
    out = {}
    for part in s.split(";"):
        part = part.strip()
        if not part or "=" not in part:
            continue
        k, v = part.split("=", 1)
        out[k.strip()] = v.strip()
    return out


# ----------------------------------------------------------------------------- #
#  ПОСТРОЕНИЕ ТЕЛА ЗАПРОСА                                                       #
# ----------------------------------------------------------------------------- #

def default_json_query(newobject_id, region_id, page, engine_version,
                       room=None, price_min=None, price_max=None, sort=None):
    """Шаблон jsonQuery по устоявшейся схеме search-offers-desktop.

    Структура подтверждена по реальным запросам (см. README): region/terms,
    _type=flatsale, engine_version/term, room/terms с кодами 1..6,9(студия),7(своб.),
    price/range {gte,lte}, page/term, sort/term. newobject/terms — фильтр по ЖК.
    """
    q = {
        "_type": "flatsale",
        "engine_version": {"type": "term", "value": engine_version},
        "region": {"type": "terms", "value": [region_id]},
        "newobject": {"type": "terms", "value": [newobject_id]},
        "page": {"type": "term", "value": page},
    }
    if sort:
        # стабильная сортировка — чтобы границы страниц не «плавали» при пагинации
        q["sort"] = {"type": "term", "value": sort}
    if room is not None:
        q["room"] = {"type": "terms", "value": [room]}
    if price_min is not None or price_max is not None:
        rng = {}
        if price_min is not None:
            rng["gte"] = price_min
        if price_max is not None:
            rng["lte"] = price_max
        q["price"] = {"type": "range", "value": rng}
    return {"jsonQuery": q}


def query_from_template(template, newobject_id, region_id, page, engine_version,
                        room=None, price_min=None, price_max=None):
    """
    Берёт тело из вашего cURL как ШАБЛОН и подставляет только фильтры выгрузки.
    Всё остальное (region, sort, флаги) сохраняется из реального запроса.
    """
    body = json.loads(json.dumps(template))   # глубокая копия
    q = body.get("jsonQuery")
    if not isinstance(q, dict):
        # тело без jsonQuery — не наш шаблон, откатываемся на дефолт
        return default_json_query(newobject_id, region_id, page, engine_version,
                                  room, price_min, price_max)
    q["_type"] = q.get("_type", "flatsale")
    q["newobject"] = {"type": "terms", "value": [newobject_id]}
    q["page"] = {"type": "term", "value": page}
    if room is not None:
        q["room"] = {"type": "terms", "value": [room]}
    else:
        q.pop("room", None)
    if price_min is not None or price_max is not None:
        rng = {}
        if price_min is not None:
            rng["gte"] = price_min
        if price_max is not None:
            rng["lte"] = price_max
        q["price"] = {"type": "range", "value": rng}
    else:
        q.pop("price", None)
    return body


# ----------------------------------------------------------------------------- #
#  СЕТЬ: сессия, ретраи, один запрос                                            #
# ----------------------------------------------------------------------------- #

class Fetcher:
    """Инкапсулирует requests.Session, заголовки/cookie, ретраи и задержки."""

    def __init__(self, args, curl=None):
        self.args = args
        self.session = requests.Session()
        headers = dict(DEFAULT_HEADERS)
        self.api_url = API_URL
        self.template = None
        if curl:
            if curl.get("headers"):
                headers.update(curl["headers"])
            if curl.get("cookies"):
                self.session.cookies.update(curl["cookies"])
            if curl.get("url"):
                self.api_url = curl["url"]
            self.template = curl.get("json")
        # --cookie "k=v; k2=v2" — быстрый способ передать сессию без полного cURL
        if getattr(args, "cookie", None):
            self.session.cookies.update(_parse_cookie_string(args.cookie))
        # --user-agent переопределение
        if getattr(args, "user_agent", None):
            headers["User-Agent"] = args.user_agent
        # Referer на страницу ЖК (антибан)
        headers.setdefault("Referer", args.referer or "https://www.cian.ru/")
        self.session.headers.update(headers)

    def build_body(self, page, room=None, price_min=None, price_max=None):
        if self.template:
            return query_from_template(
                self.template, self.args.jk_id, self.args.region, page,
                self.args.engine_version, room, price_min, price_max)
        return default_json_query(
            self.args.jk_id, self.args.region, page, self.args.engine_version,
            room, price_min, price_max, sort=getattr(self.args, "sort", None))

    def post(self, body):
        """Один POST с экспоненциальным бэк-оффом на 429/403/5xx и сетевых сбоях."""
        delay = self.args.backoff_base
        last_err = None
        for attempt in range(1, max(1, self.args.retries) + 1):
            try:
                r = self.session.post(self.api_url, json=body, timeout=self.args.timeout)
                if r.status_code == 200:
                    return r.json()
                if r.status_code in (403, 429) or r.status_code >= 500:
                    last_err = f"HTTP {r.status_code}"
                    log.warning("  %s (попытка %d/%d) — пауза %.1fs",
                                last_err, attempt, max(1, self.args.retries), delay)
                    time.sleep(delay)
                    delay *= 2
                    continue
                # прочие 4xx — постоянная ошибка, не ретраим (падаем сразу)
                raise RuntimeError(
                    f"HTTP {r.status_code} — запрос отклонён, не ретраится: {r.text[:200]}")
            except (requests.RequestException, json.JSONDecodeError) as e:
                last_err = repr(e)
                log.warning("  сетевой сбой (попытка %d/%d): %s — пауза %.1fs",
                            attempt, self.args.retries, last_err, delay)
                time.sleep(delay)
                delay *= 2
        raise RuntimeError(f"Запрос не удался после {self.args.retries} попыток: {last_err}")

    def pause(self):
        time.sleep(random.uniform(self.args.delay_min, self.args.delay_max))


# ----------------------------------------------------------------------------- #
#  ИЗВЛЕЧЕНИЕ ДАННЫХ ИЗ ОТВЕТА                                                   #
# ----------------------------------------------------------------------------- #

def get_offers_and_count(resp):
    """
    Возвращает (offers_list, total_count). Схема ответа защищена:
    offers — data.offersSerialized | data.offers | data.items[].offer;
    count  — data.offerCount | offersCount | totalCount | len(offers).
    """
    if not isinstance(resp, dict):
        return [], 0
    data = resp.get("data", resp)
    offers = (data.get("offersSerialized")
              or data.get("offers")
              or resp.get("offers")
              or data.get("items")        # альтернативный формат items[].offer
              or [])
    # формат items[].offer -> развернуть до плоских офферов
    offers = [it.get("offer", it) if isinstance(it, dict) else it for it in offers]
    count = (data.get("offerCount")
             or data.get("offersCount")
             or data.get("totalCount")
             or resp.get("offerCount")
             or len(offers))
    try:
        count = int(count)
    except (TypeError, ValueError):
        count = len(offers)
    return offers, count


def dig(obj, path, default=None):
    """Безопасный доступ по точечному пути 'a.b.c'. Любая дыра -> default."""
    cur = obj
    for key in path.split("."):
        if isinstance(cur, dict):
            cur = cur.get(key)
        else:
            return default
        if cur is None:
            return default
    return cur


# ----------------------------------------------------------------------------- #
#  СБОР: пагинация + room-split + price-split + дедуп                            #
# ----------------------------------------------------------------------------- #

def collect_segment(fetcher, room, price_min, price_max):
    """
    Листает одну выборку (room+price) до конца или до потолка страниц.
    Возвращает (offers_by_id: dict, total_count: int|None, hit_cap: bool).
    """
    by_id = {}
    total = None
    page = 1
    max_pages = fetcher.args.max_pages
    while page <= max_pages:
        body = fetcher.build_body(page, room, price_min, price_max)
        resp = fetcher.post(body)
        offers, total = get_offers_and_count(resp)
        if not offers:
            break
        for o in offers:
            cid = o.get("cianId") or o.get("id")
            if cid is not None:
                by_id[cid] = o
        seg = _seg_label(room, price_min, price_max)
        log.info("    %s стр.%d: +%d (всего собрано %d, на Циан ~%s)",
                 seg, page, len(offers), len(by_id), total)
        if len(offers) < PAGE_SIZE:
            break
        page += 1
        if page <= max_pages:
            fetcher.pause()
    hit_cap = (page > max_pages)
    return by_id, total, hit_cap


def collect_room(fetcher, room):
    """
    Собирает все лоты одной комнатной категории, при упоре в потолок дробит
    диапазон цены пополам и рекурсивно добирает. Дедуп по cianId.
    """
    by_id = {}
    room_total = None        # offerCount по этой категории (для охвата в Сводке)
    # стек диапазонов (lo, hi); None => без границы
    stack = [(fetcher.args.price_min, fetcher.args.price_max)]
    min_span = fetcher.args.min_price_span
    first_segment = True
    while stack:
        lo, hi = stack.pop()
        seg, total, hit_cap = collect_segment(fetcher, room, lo, hi)
        if first_segment:
            room_total = total   # первый сегмент — полный диапазон => общее число по категории
            first_segment = False
        by_id.update(seg)
        collected = len(seg)
        need_split = (
            fetcher.args.split_price
            and total is not None
            and total > collected            # выдача не вместилась
            and total > PAGE_SIZE * fetcher.args.max_pages  # реально упёрлись в потолок
        )
        if need_split:
            nlo = 0 if lo is None else lo
            nhi = fetcher.args.price_ceiling if hi is None else hi
            if nhi - nlo > min_span:
                mid = (nlo + nhi) // 2
                log.info("    дроблю цену [%s..%s] -> [%s..%s] + [%s..%s] (на Циан ~%s, взято %d)",
                         nlo, nhi, nlo, mid, mid + 1, nhi, total, collected)
                stack.append((nlo, mid))
                stack.append((mid + 1, nhi))
            else:
                log.warning("    диапазон [%s..%s] не дробится дальше (min-span=%s), "
                            "часть лотов может быть недобрана", nlo, nhi, min_span)
    return by_id, room_total


def probe_total(fetcher):
    """Один запрос (страница 1, без room-фильтра) — узнать offerCount по всему ЖК."""
    try:
        resp = fetcher.post(fetcher.build_body(1))
        _, total = get_offers_and_count(resp)
        return total
    except Exception as e:
        log.warning("Не удалось узнать общее число лотов в ЖК: %s", e)
        return None


def collect_all(fetcher):
    """
    Главный сбор: по умолчанию обходим категории комнатности, затем дедуп.
    Возвращает (offers, totals_by_room) — totals_by_room для охвата в Сводке.
    """
    by_id = {}
    totals_by_room = {}
    if fetcher.args.no_room_split:
        log.info("  Сбор без room-split (одним запросом по ЖК)...")
        res, _ = collect_room(fetcher, None)
        by_id.update(res)
        totals_by_room = None
    else:
        for room in ROOM_SPLIT_ORDER:
            log.info("  Категория room=%s ...", room)
            res, room_total = collect_room(fetcher, room)
            new = sum(1 for k in res if k not in by_id)
            by_id.update(res)
            if room_total is not None:
                totals_by_room[room] = room_total
            fetcher.pause()
            log.info("  room=%s: собрано %d (всего на Циан ~%s), новых для ЖК %d, итого %d",
                     room, len(res), room_total, new, len(by_id))
    return list(by_id.values()), totals_by_room


def _seg_label(room, pmin, pmax):
    parts = []
    if room is not None:
        parts.append(f"room={room}")
    if pmin is not None or pmax is not None:
        parts.append(f"₽[{pmin or 0}..{pmax or '∞'}]")
    return " ".join(parts) if parts else "ЖК"


# ----------------------------------------------------------------------------- #
#  НОРМАЛИЗАЦИЯ ОДНОГО ЛОТА                                                      #
# ----------------------------------------------------------------------------- #

def category_of(o):
    """Человекочитаемая комнатность (Студия / Своб. планировка / 1 / 2 / 3 / 4+)."""
    if o.get("isStudio") or o.get("flatType") == "studio":
        return CAT_STUDIO
    ft = o.get("flatType")
    if ft in ("openPlan", "openplan", "freePlan"):
        return CAT_OPENPLAN
    rc = o.get("roomsCount")
    if rc is None:
        # признака комнатности нет вовсе (напр. карточка из Playwright-DOM) -> неизвестно
        rc = o.get("roomsForSaleCount")
        if rc is None:
            return None
    try:
        rc = int(rc)
    except (TypeError, ValueError):
        return None
    if rc == 0:
        return CAT_STUDIO   # Циан кодирует студию как 0 комнат
    if rc >= 4:
        return "4+"
    return str(rc)


# ===== FIN-BLOCK-START ======================================================= #
#  ОТДЕЛКА/РЕМОНТ — порт блока из extension/content.js                           #
# ----------------------------------------------------------------------------- #
# Один и тот же набор правил живёт в трёх экспортёрах: extension/content.js,
# cian_browser.js и здесь. Правите один — правьте все три; расхождение ловит
# tests/check_finish.mjs (JS) и --self-test (Python) на общем корпусе
# tests/finish_corpus.json.
#
# Слои по убыванию надёжности:
#   1) поле Циан repairType/decoration -> источник «Циан-поле»
#   2) разбор текста объявления        -> источник «из описания»
#   3) не нашлось                      -> категория не определена
#
# В регулярках ниже ЗАПРЕЩЕНЫ \w, \b, \d, lookbehind и флаги i/u: в JS \w и \b
# ASCII-only и на кириллице молча не срабатывают, а в Python — срабатывают.
# Держим общий знаменатель, иначе один и тот же оффер получит РАЗНЫЕ категории
# в выгрузке из расширения и из скрипта. Регистр и «ё» снимает _fin_norm().

FIN = {
    "none": "Без отделки", "rough": "Черновая", "prefine": "Предчистовая (white box)",
    "fine": "Чистовая", "turnkey": "Под ключ / с мебелью",
    "norepair": "Без ремонта", "cosmetic": "Косметический", "euro": "Евроремонт",
    "designer": "Дизайнерский", "some": "С ремонтом (тип не указан)",
}

# Значение поля Циан -> категория. Подтверждены дампами API:
#   decoration: without | rough | preFine | fine | fineWithFurniture
#   repairType: no | cosmetic | euro | design
# Остальные ключи — толерантные догадки на случай смены словаря.
_FIELD_FIN = {
    "without": FIN["none"], "rough": FIN["rough"], "draft": FIN["rough"],
    "prefine": FIN["prefine"], "preFine": FIN["prefine"], "whitebox": FIN["prefine"],
    "fine": FIN["fine"], "clean": FIN["fine"], "finish": FIN["fine"],
    "chistovaya": FIN["fine"],
    "fineWithFurniture": FIN["turnkey"], "turnkey": FIN["turnkey"],
    "withFurniture": FIN["turnkey"],
    "no": FIN["norepair"], "norepair": FIN["norepair"],
    "cosmetic": FIN["cosmetic"], "normal": FIN["cosmetic"],
    "euro": FIN["euro"], "good": FIN["euro"],
    "design": FIN["designer"], "designer": FIN["designer"],
}

# Стоп-контексты: вырезаются из текста ДО классификации, чтобы ремонт подъезда,
# соседнего корпуса или «сделаем под ваш вкус» не приписывался самой квартире.
# Вырезается только найденный участок, а не всё предложение.
_FIN_STOPS = [
    re.compile(r"(?:^|[^а-яё])(?:кап(?:итальн[а-яё]*)?[\s-]*)?(?:ремонт|отделк)[а-яё]*[\s-]*(?:в[\s-]*|на[\s-]*)?(?:детск[а-яё]*[\s-]*(?:сад|площадк)|мест[а-яё]*[\s-]*общего|подъезд|фасад|кровл|крыш|дорог|тротуар|лифт|двор|подвал|чердак|площадк|набережн|станц|метро|улиц|шоссе|проспект|школ|моп|стояк|трубопровод|инженерн|паркинг|парковк|холл|лобби|вестибюл|входн[а-яё]*[\s-]*групп)[а-яё]*"),   # S1 — ремонт общедомового/городского объекта: подъезд, фасад, дорога, лифт
    re.compile(r"(?:^|[^а-яё])(?:кап(?:итальн[а-яё]*)?[\s-]*)?(?:ремонт|отделк)[а-яё]*[\s-]*(?:в[\s-]*)?(?:дом[аеу]|здани|корпус|многоквартирн)[а-яё]*"),   # S2 — ремонт дома/здания/корпуса целиком
    re.compile(r"(?:^|[^а-яё])(?:дом|здани|корпус|подъезд|фасад|кровл|крыш|школ|поликлиник|детск[а-яё]*[\s-]*сад)[а-яё]*[\s-]*(?:был[а-яё]*[\s-]*|уже[\s-]*|недавно[\s-]*)?(?:после|прошел|прошла|прошли|ждет|ожидает|планируется|стоит[\s-]*в[\s-]*плане|под)[\s-]*(?:кап(?:итальн[а-яё]*)?[\s-]*)?(?:ремонт|отделк)[а-яё]*"),   # S3 — обратный порядок: «дом после капремонта»
    re.compile(r"(?:^|[^а-яё])(?:в|во)[\s-]*(?:дом[еу]|подъезде|здании|корпусе|дворе|холле|лобби|местах[\s-]*общего[\s-]*пользования)(?![а-яё])(?:(?!квартир|апартамент|комнат)[^.!?;,]){0,12}?(?:ремонт|отделк)[а-яё]*|(?:^|[^а-яё])(?:в|во)[\s-]*(?:дом[еу]|подъезде|здании|корпусе|дворе|холле|лобби|местах[\s-]*общего[\s-]*пользования)(?![а-яё])[\s-]*(?:(?:уже|недавно|полностью|сейчас|как[\s-]*раз|только[\s-]*что)[\s-]*)*(?:сделан|выполнен|проведен|проведён|завершен|завершён|идет|идёт|ведется|ведётся|планируется|запланирован)[а-яё]*[\s-]*(?:кап(?:итальн[а-яё]*)?[\s-]*)?(?:ремонт|отделк)[а-яё]*"),   # S4 — локатив переносит ремонт на дом: «в доме сделан ремонт»
    re.compile(r"(?:^|[^а-яё])(?:ремонт|отделк)[а-яё]*[\s-]*(?:в|у)[\s-]*(?:соседн|друг|перв|втор|треть|остальн)[а-яё]*[\s-]*(?:корпус|дом|подъезд|квартир|секц|башн|блок|очеред)[а-яё]*"),   # S5 — чужой объект: «ремонт в соседнем корпусе»
    re.compile(r"(?:^|[^а-яё])(?:в|во|у)[\s-]*(?:соседн[а-яё]*|сосед[а-яё]*|друг[а-яё]*)[\s-]*(?:корпус|дом|подъезд|секц|башн|блок|очеред)?[а-яё]*(?:(?!квартир|апартамент|комнат)[^.!?;,]){0,30}?(?:ремонт|отделк)[а-яё]*"),   # S6 — чужой объект, обратный порядок: «у соседей евроремонт»
    re.compile(r"(?:^|[^а-яё])(?:рядом|неподалеку|поблизости|напротив|через[\s-]*дорогу|по[\s-]*соседству|во[\s-]*дворе)(?:(?!квартир|апартамент|комнат)[^.!?;,]){0,30}?(?:ремонт|отделк)[а-яё]*"),   # S7 — окружение, а не лот: «рядом идёт ремонт»
    re.compile(r"(?:^|[^а-яё])(?:сделаем|сделаю|выполним|поможем|организуем|подберем|обеспечим|доделаем|предлагаем|обсуждаем|можем[\s-]*сделать|можно[\s-]*(?:сделать|заказать)|готовы[\s-]*(?:сделать|выполнить)|возможн[а-яё]*|планируетс[а-яё]*|остал[а-яё]*[\s-]*(?:сделать|доделать))(?:(?!квартир|апартамент|комнат)[^.!?;,]){0,30}?(?:ремонт|отделк)[а-яё]*(?:[\s-]*(?:под[\s-]*ключ|от[\s-]*застройщика|под[\s-]*ваш[а-яё]*[\s-]*вкус|за[\s-]*доплату))*"),   # S8 — будущий/гипотетический ремонт: «сделаем ремонт под ваш вкус»
    re.compile(r"(?:^|[^а-яё])(?:ремонт|отделк)[а-яё]*(?:(?!квартир|апартамент|комнат)[^.!?;,]){0,30}?(?:за[\s-]*доплату|под[\s-]*ваш[а-яё]*[\s-]*вкус|по[\s-]*ваш[а-яё]*[\s-]*проект[а-яё]*|под[\s-]*заказ|по[\s-]*желани[а-яё]*[\s-]*покупател[а-яё]*|на[\s-]*ваш[\s-]*выбор|опционально)"),   # S9 — опциональность: «отделка за доплату»
    re.compile(r"(?:^|[^а-яё])(?:ремонт|отделк)[а-яё]*[\s-]*(?:в[\s-]*подарок|в[\s-]*кредит|в[\s-]*рассрочку|в[\s-]*ипотеку|за[\s-]*счет[\s-]*(?:банка|застройщика))"),   # S10 — ремонт как бонус/финпродукт: «ремонт в подарок»
    re.compile(r"(?:^|[^а-яё])(?:скидка|рассрочк|кредит|ипотек|субсиди|бонус|сертификат|смет|материал|бригад|подрядчик|дизайн[\s-]*студи)[а-яё]*(?:(?!квартир|апартамент|комнат)[^.!?;,]){0,30}?(?:на[\s-]*)?(?:ремонт|отделк)[а-яё]*"),   # S11 — реклама услуг и финансирования ремонта: «рассрочка на ремонт»
]

# Порядок = приоритет: явные категории раньше общих, отрицание раньше
# утверждения («не требует ремонта» опережает «требует ремонта»), качество
# отделки важнее меблировки. Последнее правило — catch-all.
_FIN_RULES = [
    (FIN["designer"], re.compile(r"(?:^|[^а-яё])дизайнерск[а-яё]*[\s-]*(?:ремонт|отделк|интерьер|квартир|апартамент|решени|проект)[а-яё]*|(?:^|[^а-яё])(?:авторск|эксклюзивн)[а-яё]*[\s-]*(?:ремонт|отделк|интерьер|проект)[а-яё]*|(?:ремонт|отделк[а-яё]*|интерьер[а-яё]*)[\s-]*(?:(?:полностью|целиком)[\s-]*)?(?:выполнен[а-яё]*|сделан[а-яё]*|разработан[а-яё]*)?[\s-]*по[\s-]*(?:(?:индивидуальн|авторск|специальн)[а-яё]*[\s-]*)*дизайн[\s-]*проект[а-яё]*|(?:ремонт|отделк[а-яё]*|интерьер[а-яё]*)[\s-]*от[\s-]*(?:известн[а-яё]*[\s-]*)?дизайнер[а-яё]*|(?:^|[^а-яё])реализован[а-яё]*[\s-]*дизайн[\s-]*проект[а-яё]*|(?:^|[^а-яё])(?:отделк|ремонт|интерьер)[а-яё]*[\s:-]*(?:выполнен[а-яё]*[\s:-]*)?дизайнерск[а-яё]*")),
    (FIN["euro"], re.compile(r"(?:^|[^а-яё])евро[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])евро[\s-]*отделк[а-яё]*|(?:^|[^а-яё])евростандарт[а-яё]*|ремонт[а-яё]*[\s-]*в[\s-]*евро[\s-]*стиле|(?:^|[^а-яё])(?:отделк|ремонт)[а-яё]*[\s:-]*евро[а-яё]*")),
    (FIN["prefine"], re.compile(r"white[\s-]*box|(?:^|[^а-яё])(?:вайт|уайт)[\s-]*бокс[а-яё]*|(?:^|[^а-яё])пред[\s-]*чистов[а-яё]*|(?:^|[^а-яё])под[\s-]*чистов[а-яё]*|(?:^|[^а-яё])улучшенн[а-яё]*[\s-]*чернов[а-яё]*|(?:^|[^а-яё])(?:отделк|ремонт)[а-яё]*[\s:-]*(?:предчистов[а-яё]*|white[\s-]*box)")),
    (FIN["rough"], re.compile(r"(?:^|[^а-яё])чернов[а-яё]*[\s-]*(?:отделк|состоян|вариант|вид)[а-яё]*|(?:^|[^а-яё])чернов(?:ая|ой|ую|ое)(?![а-яё])|(?:^|[^а-яё])(?:с|со)[\s-]*чернов[а-яё]*|(?:^|[^а-яё])(?:отделк|ремонт)[а-яё]*[\s:-]*чернов[а-яё]*")),
    (FIN["none"], re.compile(r"(?:^|[^а-яё])без[\s-]*(?:как[а-яё]*[\s-]*либо[\s-]*|всяк[а-яё]*[\s-]*)?(?:[а-яё]+(?:ой|ей|ий|ый|ая|ое|ых|ым)[\s-]+){0,2}отделк[а-яё]*|(?:^|[^а-яё])нет[\s-]*отделк[а-яё]*|отделк[а-яё]*[\s-]*(?:полностью[\s-]*)?отсутству[а-яё]*|(?:^|[^а-яё])не[\s-]*выполнен[а-яё]*[\s-]*отделк[а-яё]*|отделк[а-яё]*[\s-]*не[\s-]*(?:выполнен|сделан|производ)[а-яё]*|(?:^|[^а-яё])голы[ех][\s-]*стен[а-яё]*|(?:^|[^а-яё])бетонн[а-яё]*[\s-]*коробк[а-яё]*|отделк[а-яё]*[\s-]*не[\s-]*предусмотрен[а-яё]*|(?:^|[^а-яё])отделк[а-яё]*[\s-]*нет(?![а-яё])")),
    (FIN["fine"], re.compile(r"(?:^|[^а-яё])чистов[а-яё]*[\s-]*отделк[а-яё]*|отделк[а-яё]*[\s-]*(?:от[\s-]*)?застройщик[а-яё]*|(?:^|[^а-яё])(?:с|со)[\s-]*(?:полной[\s-]*|готовой[\s-]*|качественной[\s-]*|финишной[\s-]*|чистовой[\s-]*)?отделк[а-яё]*|(?:^|[^а-яё])готов[а-яё]*[\s-]*отделк[а-яё]*|отделк[а-яё]*[\s-]*(?:уже[\s-]*)?(?:выполнен|сделан|готов)[а-яё]*|(?:^|[^а-яё])сдан[а-яё]*[\s-]*(?:с|со)[\s-]*отделк[а-яё]*|(?:^|[^а-яё])(?:отделк|ремонт)[а-яё]*[\s:-]*чистов[а-яё]*")),
    (FIN["some"], re.compile(r"(?:^|[^а-яё])не[\s-]*требу[а-яё]*[\s-]*(?:[а-яё]+[\s-]+){0,2}(?:ремонт|вложен|отделк)[а-яё]*|(?:^|[^а-яё])ремонт[а-яё]*[\s-]*(?:[а-яё]+[\s-]+){0,2}не[\s-]*требу[а-яё]*|(?:^|[^а-яё])не[\s-]*нужен[\s-]*ремонт[а-яё]*")),
    (FIN["norepair"], re.compile(r"(?:^|[^а-яё])без[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])требу[а-яё]*[\s-]*(?:[а-яё]+[\s-]+){0,2}ремонт[а-яё]*|(?:^|[^а-яё])нужен[\s-]*(?:[а-яё]+[\s-]+){0,1}ремонт[а-яё]*|(?:^|[^а-яё])нужда[а-яё]*[\s-]*в[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])под[\s-]*ремонт(?![а-яё])|(?:^|[^а-яё])убит[а-яё]*[\s-]*(?:квартир|состоян|двушк|трешк|однушк)[а-яё]*|(?:^|[^а-яё])(?:в|во)[\s-]*(?:строительн|первоначальн|плачевн|ужасн|убит|предремонтн)[а-яё]*[\s-]*состоян[а-яё]*|(?:^|[^а-яё])ремонт[а-яё]*[\s-]*(?:[а-яё]+[\s-]+){0,2}не[\s-]*(?:было|делал|начат|производ|провод|дела)[а-яё]*|(?:^|[^а-яё])(?:никогда[\s-]*)?не[\s-]*(?:делал|производил|проводил)[а-яё]*[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])(?:бабушкин|дедушкин|советск)[а-яё]*[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])требует[\s-]*вложени[а-яё]*")),
    (FIN["cosmetic"], re.compile(r"(?:^|[^а-яё])косметич[а-яё]*|(?:^|[^а-яё])космет(?![а-яё])[\s-]*ремонт|(?:^|[^а-яё])(?:в[\s-]*)?(?:жило[а-яё]*|хорош[а-яё]*|отличн[а-яё]*|нормальн[а-яё]*|приличн[а-яё]*|достойн[а-яё]*|ухожен[а-яё]*)[\s-]*состоян[а-яё]*|(?:^|[^а-яё])(?:сделан|выполнен|произведен|проведен)[а-яё]*[\s-]*(?:[а-яё]+[\s-]+){0,2}ремонт[а-яё]*|(?:^|[^а-яё])после[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])(?:свеж|недавн|нов|аккуратн|добротн|качественн|современн|легк|хорош|отличн|приличн|достойн)[а-яё]*[\s-]*ремонт[а-яё]*|(?:^|[^а-яё])ремонт[а-яё]*[\s-]*(?:сделан|выполнен)[а-яё]*|(?:^|[^а-яё])(?:отделк|ремонт)[а-яё]*[\s:-]*косметическ[а-яё]*")),
    (FIN["turnkey"], re.compile(r"(?:^|[^а-яё])под[\s-]*ключ(?![а-яё])|(?:^|[^а-яё])(?:с|со)[\s-]*(?:всей[\s-]*|полной[\s-]*|новой[\s-]*)?мебел[а-яё]*|(?:^|[^а-яё])меблирован[а-яё]*|(?:^|[^а-яё])(?:с|со)[\s-]*(?:быт[а-яё]*[\s-]*)?техник(?:а|и|е|у|ой)?(?![а-яё])|(?:^|[^а-яё])(?:вся[\s-]*)?мебел[а-яё]*[\s-]*(?:и[\s-]*техник[а-яё]*[\s-]*)?оста(?:ет|ю)[а-яё]*|(?:^|[^а-яё])оста(?:ет|ю)[а-яё]*[\s-]*(?:вся[\s-]*)?мебел[а-яё]*|(?:^|[^а-яё])полностью[\s-]*обставлен[а-яё]*")),
    (FIN["some"], re.compile(r"(?:^|[^а-яё])(?:можно[\s-]*(?:сразу[\s-]*)?(?:жить|заезжать|въезжать|заселяться)|(?:за|в)езжай[\s-]*и[\s-]*живи|готов[а-яё]*[\s-]*к[\s-]*(?:заселени|проживани)[а-яё]*)")),
    (FIN["some"], re.compile(r"(?:^|[^а-яё])ремонт[а-яё]*|(?:^|[^а-яё])отремонтирован[а-яё]*|(?:^|[^а-яё])(?:с|со)[\s-]*отделк[а-яё]*")),
]

DESC_MAX_CHARS = 600        # предел читаемости листа, не предел Excel
EXCEL_CELL_LIMIT = 32767    # жёсткий предел длины строки в ячейке
_TAGS = re.compile(r"<[^>]+>")
# Символы, которые openpyxl отказывается писать: присваивание строки с \x0b
# роняет IllegalCharacterError и убивает запись ВСЕЙ книги.
_ILLEGAL = re.compile(r"[\000-\010\013\014\016-\037]")
# Одинокие суррогаты приезжают из ответа API как есть (json.loads их пропускает).
# openpyxl запишет такую строку молча, а вот ОТКРЫТЬ книгу потом не получится.
_SURROGATE = re.compile("[\ud800-\udfff]")
_NBSP = re.compile("[\u00a0\u202f]")
_FIN_ZW = re.compile("[\u00ad\u200b\u200c\u200d\ufeff]")      # мягкий перенос, zero-width
_ASTRAL = re.compile("[^\u0000-\uffff]")                        # эмодзи и прочее вне BMP
# Мини-декодер HTML-сущностей — намеренно НЕ html.unescape(): полного аналога в JS
# нет, и один и тот же оффер получал бы разные категории в разных выгрузках.
_ENT_RX = re.compile(r"&(#[0-9]{1,7}|#[xX][0-9a-fA-F]{1,6}|[a-zA-Z][a-zA-Z0-9]{1,10});")
_ENT = {
    "amp": "&", "lt": "<", "gt": ">", "quot": '"', "apos": "'", "nbsp": " ", "shy": "",
    "laquo": "\u00ab", "raquo": "\u00bb", "mdash": "\u2014", "ndash": "\u2013",
    "hellip": "\u2026", "middot": "\u00b7", "times": "\u00d7", "deg": "\u00b0",
    "lsquo": "\u2018", "rsquo": "\u2019", "ldquo": "\u201c", "rdquo": "\u201d",
    "copy": "\u00a9", "reg": "\u00ae", "euro": "\u20ac", "rouble": "\u20bd",
}


def _fin_entities(s):
    def rep(m):
        e = m.group(1)
        if e[0] == "#":
            hexa = e[1] in "xX"
            try:
                cp = int(e[2:], 16) if hexa else int(e[1:], 10)
            except ValueError:
                return m.group(0)
            if not (1 <= cp <= 0x10FFFF) or 0xD800 <= cp <= 0xDFFF:
                return m.group(0)
            return chr(cp)
        v = _ENT.get(e.lower())
        return m.group(0) if v is None else v
    return _ENT_RX.sub(rep, s)


def _field_value(v):
    """Значение поля отделки: разворачиваем dict на уровень и приводим к строке.
    Сложное значение считаем отсутствующим — иначе поиск по словарю падает
    TypeError на нехешируемом ключе и роняет весь прогон."""
    if isinstance(v, dict):
        v = v.get("type") or v.get("value")
    if v is None or isinstance(v, (list, dict, set, tuple)):
        return None
    if isinstance(v, str):
        return v.strip() or None
    return str(v)
_FIN_DASH = re.compile("[\u2010-\u2015\u2212]")                 # типографские тире
# Явный класс пробелов вместо \s — см. комментарий в JS-версии блока.
_FIN_WS = re.compile("[\u0009-\u000d\u001c-\u001f\u0020\u0085\u00a0\u1680"
                     "\u2000-\u200a\u2028\u2029\u202f\u205f\u3000]+")


def description_of(o):
    """Полный текст объявления, вычищенный до пригодного для Excel. None, если пусто.

    Подтверждённое поле Циан — только `description` (строка на верхнем уровне
    оффера). `title` у квартир обычно пустой, но у карточек из playwright-фолбэка
    это единственный доступный текст, поэтому он идёт запасным вариантом.
    НЕ обрезается: обрезка — дело _clip_desc(), а классификация обязана видеть
    текст целиком.
    """
    d = o.get("description")
    if isinstance(d, dict):
        d = d.get("text") or d.get("value")
    if not isinstance(d, str) or not d.strip():
        t = o.get("title")
        d = t if isinstance(t, str) else ""
    if not d:
        return None
    d = _TAGS.sub(" ", d)
    d = _fin_entities(d)
    d = _ILLEGAL.sub("", d)
    d = _SURROGATE.sub("", d)
    d = _FIN_ZW.sub("", d)
    d = _NBSP.sub(" ", d)
    d = _FIN_WS.sub(" ", d).strip(" ")
    if not d:
        return None
    return d[:EXCEL_CELL_LIMIT]


def _clip_desc(d):
    """Обрезка ТОЛЬКО для ячейки: 600 знаков — предел читаемости листа."""
    if not d:
        return d
    return d if len(d) <= DESC_MAX_CHARS else d[:DESC_MAX_CHARS] + "\u2026"


def _fin_norm(t):
    s = ("" if t is None else str(t)).lower().replace("\u0451", "\u0435")   # ё -> е
    # Эмодзи и одинокие суррогаты: в JS это ДВА code unit, в Python — один символ,
    # а окна {0,n} в стоп-контекстах считают единицы движка. Без выпиливания один
    # и тот же текст давал бы разные категории в JS и в Python.
    s = _ASTRAL.sub(" ", s)
    s = _SURROGATE.sub(" ", s)
    s = _FIN_ZW.sub("", s)
    s = _FIN_DASH.sub("-", s)
    return _FIN_WS.sub(" ", s).strip(" ")


def finish_from_text(t):
    """Категория отделки по тексту объявления. None, если признаков нет."""
    if not t:
        return None
    s = _fin_norm(t)
    for rx in _FIN_STOPS:
        s = rx.sub(" ", s)
    s = _FIN_WS.sub(" ", s)
    for label, rx in _FIN_RULES:
        if rx.search(s):
            return label
    return None


def finish_of(o, desc=None):
    """(категория, источник). Источник: «Циан-поле» / «из описания» / None."""
    rt = _field_value(o.get("repairType"))
    dc = _field_value(o.get("decoration"))
    if rt and _FIELD_FIN.get(rt):
        return _FIELD_FIN[rt], "Циан-поле"
    if dc and _FIELD_FIN.get(dc):
        return _FIELD_FIN[dc], "Циан-поле"
    ft = finish_from_text(desc if desc is not None else description_of(o))
    if ft:
        return ft, "из описания"
    if rt or dc:
        return str(rt or dc), "Циан-поле"      # словарь отстал от Циан
    return None, None


# ===== FIN-BLOCK-END ========================================================= #


def seller_type_of(o):
    """Возвращает 'Застройщик' / 'Собственник' / 'Агентство' / None."""
    if (o.get("isFromBuilder") or o.get("fromDeveloper")
            or dig(o, "newbuilding.isFromBuilder")
            or dig(o, "newbuilding.isFromDeveloper")
            or dig(o, "user.isDeveloper")
            or dig(o, "user.userType") in ("developer", "builder")):
        return "Застройщик"
    if o.get("isByHomeowner") or dig(o, "user.userType") in ("homeowner", "owner"):
        return "Собственник"
    ut = dig(o, "user.userType")
    if ut in ("agency", "realtor", "agent", "managementCompany"):
        return "Агентство"
    # эвристика: есть имя агентства -> агентство, иначе неизвестно
    if dig(o, "user.agencyName") or dig(o, "user.companyName"):
        return "Агентство"
    return None


def seller_name_of(o):
    return (dig(o, "user.agencyName")
            or dig(o, "user.companyName")
            or dig(o, "user.title")
            or dig(o, "user.name")
            or None)


def building_of(o):
    return (dig(o, "newbuilding.house.name")
            or dig(o, "newbuilding.name")
            or dig(o, "building.name")
            or dig(o, "geo.jk.house.name")
            or dig(o, "house.name")
            or None)


def price_of(o):
    p = (dig(o, "bargainTerms.priceRur")
         or dig(o, "bargainTerms.price")
         or dig(o, "bargainTerms.prices.rur")
         or o.get("price"))
    try:
        return float(p) if p is not None else None
    except (TypeError, ValueError):
        return None


def area_of(o):
    a = o.get("totalArea") or dig(o, "areaParts.0.area")
    try:
        return float(str(a).replace(",", ".")) if a is not None else None
    except (TypeError, ValueError):
        return None


def _ts_to_date(ts):
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).date()
    except (TypeError, ValueError, OSError):
        return None


def _iso_to_date(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00")).date()
    except (TypeError, ValueError):
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", str(s))
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        return None


def published_date_of(o):
    ts = o.get("addedTimestamp") or o.get("creationTimestamp")
    d = _ts_to_date(ts)
    if d:
        return d
    return _iso_to_date(o.get("creationDate") or o.get("added"))


def updated_date_of(o):
    return _iso_to_date(o.get("editDate") or o.get("updatedAt")
                        or dig(o, "humanizedTimedelta"))


def offer_url_of(o):
    url = o.get("fullUrl")
    if url:
        return url
    cid = o.get("cianId") or o.get("id")
    if cid:
        return f"https://www.cian.ru/sale/flat/{cid}/"
    return None


def normalize(o, today):
    """Преобразует сырой оффер в плоский dict под колонки Excel. Поля-дыры -> None."""
    cid = o.get("cianId") or o.get("id")
    area = area_of(o)
    price = price_of(o)
    ppm = round(price / area) if (price and area) else None
    pub = published_date_of(o)
    exposure = (today - pub).days if pub else None
    floor = o.get("floorNumber")
    floors = dig(o, "building.floorsCount") or dig(o, "floorsCount")
    desc = description_of(o)                 # полный текст, один разбор на лот
    fin, fin_src = finish_of(o, desc)        # категория — по ПОЛНОМУ тексту
    return {
        "cianId": cid,
        "url": offer_url_of(o),
        "category": category_of(o),
        "area": area,
        "floor": floor,
        "floors": floors,
        "building": building_of(o),
        "seller_type": seller_type_of(o),
        "seller_name": seller_name_of(o),
        "decoration": fin,
        "finish_src": fin_src,
        "description": _clip_desc(desc),   # в ячейку — обрезанный
        "price": price,
        "ppm": ppm,
        "published": pub,
        "exposure_days": exposure,
        "updated": updated_date_of(o),
        "in_calc": DUPLICATE_FLAG_IN,
    }


# ----------------------------------------------------------------------------- #
#  EXCEL                                                                         #
# ----------------------------------------------------------------------------- #

# (заголовок, ключ нормализованного лота, ширина, числовой формат)
COLUMNS = [
    ("№",                       "_idx",          5,  "0"),
    ("ID объявления",           "cianId",        13, "0"),
    ("Категория",               "category",      11, None),
    ("Площадь, м²",             "area",          11, "0.0"),
    ("Этаж",                    "floor",         7,  "0"),
    ("Этаж-ность",              "floors",        9,  "0"),
    ("Корпус / секция",         "building",      16, None),
    ("Тип продавца",            "seller_type",   13, None),
    ("Продавец",                "seller_name",   22, None),
    ("Отделка/ремонт",          "decoration",    24, None),
    ("Источник отделки",        "finish_src",    16, None),
    ("Цена, ₽",                 "price",         15, "#,##0"),
    ("Цена за м², ₽",           "ppm",           13, "#,##0"),      # формула =price/area
    ("Дата публикации",         "published",     15, "dd.mm.yyyy"),
    ("Срок экспоз., дней",      "exposure_days", 12, "0"),          # дни с ПОСЛЕДНЕЙ подачи
    ("Дата обновления",         "updated",       15, "dd.mm.yyyy"),
    ("Описание",                "description",   46, None),
    ("Ссылка",                  "url",           10, None),
    ("В расчёте",               "in_calc",       9,  "0"),
]


def _import_openpyxl():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter


def write_workbook(rows, jk_id, jk_name, totals_by_room, out_path, total_in_jk=None):
    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter = _import_openpyxl()

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True   # Excel/LibreOffice пересчитают формулы при открытии
    today_str = date.today().strftime("%d.%m.%Y")

    header_fill = PatternFill("solid", fgColor="1F2A44")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=13)
    sub_font = Font(italic=True, color="555555", size=9)
    link_font = Font(color="1155CC", underline="single")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def col_letter(key):
        for i, (_, k, _, _) in enumerate(COLUMNS, start=1):
            if k == key:
                return get_column_letter(i)
        return None

    def write_table(ws, title, subtitle, table_rows):
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A2"] = subtitle
        ws["A2"].font = sub_font
        # шапка в строке 4
        hdr_row = 4
        for ci, (label, key, width, fmt) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=hdr_row, column=ci, value=label)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = width
        # данные с строки 5
        desc_align = Alignment(horizontal="left", vertical="top", wrap_text=False)
        area_letter = col_letter("area")
        price_letter = col_letter("price")
        for ri, row in enumerate(table_rows, start=hdr_row + 1):
            row = dict(row)
            row["_idx"] = ri - hdr_row
            for ci, (label, key, width, fmt) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=ri, column=ci)
                if key == "url":
                    url = row.get("url")
                    cell.value = "Циан →" if url else None
                    if url:
                        cell.hyperlink = url
                        cell.font = link_font
                elif key == "description":
                    desc = row.get("description")
                    # data_type='s' обязателен: текст, начинающийся с «=», openpyxl
                    # запишет как формулу, и Excel сочтёт книгу повреждённой
                    cell.value = desc or None
                    if desc:
                        cell.data_type = "s"
                    cell.alignment = desc_align
                elif key == "ppm":
                    # живая формула =цена/площадь (как в эталоне), формат — целое
                    if row.get("price") is not None and row.get("area"):
                        cell.value = f"={price_letter}{ri}/{area_letter}{ri}"
                    else:
                        cell.value = row.get("ppm")
                    cell.number_format = fmt
                else:
                    val = row.get(key)
                    cell.value = val
                    if fmt and val is not None:
                        cell.number_format = fmt
                cell.border = border
        last = hdr_row + len(table_rows)
        ws.freeze_panes = "C5"
        if table_rows:
            ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(COLUMNS))}{last}"
        return last

    # ---- Сводка ---------------------------------------------------------- #
    summary = wb.active
    summary.title = "Сводка"

    # ---- Все_лоты -------------------------------------------------------- #
    all_ws = wb.create_sheet("Все_лоты")
    uniq = len({r["cianId"] for r in rows if r.get("cianId") is not None})
    sub_all = (f"Источник: Циан (ID {jk_id}), сбор {today_str}. Собрано {len(rows)} "
               f"(уник. {uniq}). Цена за м² — живая формула =цена/площадь. "
               f"Срок экспозиции — дни с ПОСЛЕДНЕЙ подачи (Циан сбрасывает дату при переподаче).")
    last_all = write_table(all_ws, f"ЖК {jk_name} — все собранные лоты", sub_all, rows)

    # ---- Листы по комнатности ------------------------------------------- #
    cat_order = ["Студия", "Своб. планировка", "1", "2", "3", "4+"]
    cat_sheetname = {"Студия": "Студия", "Своб. планировка": "Своб_планировка",
                     "1": "1-комн", "2": "2-комн", "3": "3-комн", "4+": "4-комн"}
    present = [c for c in cat_order if any(r.get("category") == c for r in rows)]
    for cat in present:
        crows = [r for r in rows if r.get("category") == cat]
        ws = wb.create_sheet(cat_sheetname[cat])
        sub = (f"Собрано строк: {len(crows)}. Сортировка — по цене за м² "
               f"(сначала дешевле).")
        write_table(ws, f"ЖК {jk_name} — {cat}", sub, crows)

    # ---- наполняем Сводку формулами по листу Все_лоты -------------------- #
    # порядок категорий отделки берём из FIN, а не перепечатываем подписи руками:
    # иначе при первой же правке подписи строка в Сводке молча обнулится.
    # Хвостом добавляем значения, которых в FIN нет (Циан ввёл новое) — иначе они
    # утекли бы в «Не определена», хотя определены полем.
    seen_fins = {r.get("decoration") for r in rows if r.get("decoration")}
    present_fins = [lab for lab in FIN.values() if lab in seen_fins]
    present_fins += sorted(seen_fins - set(FIN.values()))
    _fill_summary(summary, all_ws.title, last_all, jk_id, jk_name, today_str,
                  totals_by_room, present, col_letter,
                  Font, Alignment, total_in_jk, present_fins)

    # порядок листов: Сводка первым
    wb.move_sheet("Сводка", -(len(wb.sheetnames) - 1))
    wb.save(out_path)


def _fill_summary(ws, sheet, last_row, jk_id, jk_name, today_str,
                  totals_by_room, present_cats, col_letter, Font, Alignment,
                  total_in_jk=None, present_fins=()):
    """Сводка: охват + средняя ₽/м² (частник vs застройщик) + диапазоны цен."""
    title_font = Font(bold=True, size=13)
    sub_font = Font(italic=True, color="555555", size=9)
    h_font = Font(bold=True)

    cat_col = col_letter("category")
    seller_col = col_letter("seller_type")
    price_col = col_letter("price")
    ppm_col = col_letter("ppm")
    calc_col = col_letter("in_calc")

    def R(col):
        return f"{sheet}!${col}$5:${col}${last_row}"

    ws["A1"] = f"ЖК {jk_name} (ID {jk_id}) — сводка по ценам"
    ws["A1"].font = title_font
    ws["A2"] = (f"Данные Циан на {today_str}. Средние, диапазоны и охват — живые формулы "
                f"по листу «{sheet}». «Частник» = собственник/агентство (вторичка/переуступка), "
                f"«Застройщик» = прямые продажи.")
    ws["A2"].font = sub_font

    # — охват выгрузки —
    r = 4
    ws.cell(r, 1, "ОХВАТ ВЫГРУЗКИ").font = h_font
    r += 1
    for i, h in enumerate(["Категория", "Собрано (уник.)", "Всего на Циан", "% выдачи"], start=1):
        ws.cell(r, i, h).font = h_font
    r += 1
    first_data = r
    for cat in present_cats:
        ws.cell(r, 1, cat)
        ws.cell(r, 2).value = f'=COUNTIFS({R(cat_col)},"{cat}",{R(calc_col)},1)'
        total = _room_total_for_cat(cat, totals_by_room)
        if total is not None:
            ws.cell(r, 3, total)
            ws.cell(r, 4).value = f"=IFERROR(B{r}/C{r},\"—\")"
            ws.cell(r, 4).number_format = "0%"
        r += 1
    ws.cell(r, 1, "ИТОГО (категории)").font = h_font
    if r > first_data:
        ws.cell(r, 2).value = f"=SUM(B{first_data}:B{r-1})"
        ws.cell(r, 3).value = f"=SUM(C{first_data}:C{r-1})"
    else:
        # ни одной категории (так бывает на пути --playwright, где нет roomsCount):
        # SUM(B6:B5) сослался бы на собственную ячейку и дал циклическую ссылку
        ws.cell(r, 2).value = 0
        ws.cell(r, 3).value = 0
    ws.cell(r, 4).value = f"=IFERROR(B{r}/C{r},\"—\")"
    ws.cell(r, 4).number_format = "0%"
    r += 1
    if total_in_jk:
        # всего активных лотов в ЖК по offerCount API (контроль полноты выгрузки)
        ws.cell(r, 1, "Всего квартир в ЖК (Циан)").font = h_font
        ws.cell(r, 2).value = f"=COUNTIFS({R(calc_col)},1)"
        ws.cell(r, 3, total_in_jk)
        ws.cell(r, 4).value = f"=IFERROR(B{r}/C{r},\"—\")"
        ws.cell(r, 4).number_format = "0%"
        r += 1
    r += 1

    # — средняя ₽/м²: частник vs застройщик —
    ws.cell(r, 1, "СРЕДНЯЯ ЦЕНА ЗА м² — ЧАСТНИК vs ЗАСТРОЙЩИК").font = h_font
    r += 1
    for i, h in enumerate(["Категория", "Частник, ₽/м²", "Застройщик, ₽/м²", "Все, ₽/м²"], start=1):
        ws.cell(r, i, h).font = h_font
    r += 1
    for cat in present_cats + ["ИТОГО по ЖК"]:
        ws.cell(r, 1, cat)
        if cat == "ИТОГО по ЖК":
            base = f'AVERAGEIFS({R(ppm_col)},{R(calc_col)},1'
            ws.cell(r, 2).value = f'=IFERROR(AVERAGEIFS({R(ppm_col)},{R(seller_col)},"<>Застройщик",{R(calc_col)},1),"—")'
            ws.cell(r, 3).value = f'=IFERROR(AVERAGEIFS({R(ppm_col)},{R(seller_col)},"Застройщик",{R(calc_col)},1),"—")'
            ws.cell(r, 4).value = f'=IFERROR({base}),"—")'
        else:
            ws.cell(r, 2).value = (f'=IFERROR(AVERAGEIFS({R(ppm_col)},{R(cat_col)},"{cat}",'
                                   f'{R(seller_col)},"<>Застройщик",{R(calc_col)},1),"—")')
            ws.cell(r, 3).value = (f'=IFERROR(AVERAGEIFS({R(ppm_col)},{R(cat_col)},"{cat}",'
                                   f'{R(seller_col)},"Застройщик",{R(calc_col)},1),"—")')
            ws.cell(r, 4).value = (f'=IFERROR(AVERAGEIFS({R(ppm_col)},{R(cat_col)},"{cat}",'
                                   f'{R(calc_col)},1),"—")')
        for cc in (2, 3, 4):
            ws.cell(r, cc).number_format = "#,##0"
        r += 1
    r += 1

    # — отделка/ремонт —
    # «Источник» здесь не украшение: он показывает, на что опирается категория.
    # Если почти всё пришло «из описания» — стоит выборочно свериться с текстом.
    fin_col = col_letter("decoration")
    src_col = col_letter("finish_src")
    ws.cell(r, 1, "ОТДЕЛКА / РЕМОНТ").font = h_font
    r += 1
    for i, h in enumerate(["Категория отделки", "Лотов", "Доля"], start=1):
        ws.cell(r, i, h).font = h_font
    r += 1
    fin_first = r
    for lab in present_fins:
        ws.cell(r, 1, lab)
        ws.cell(r, 2).value = f'=COUNTIFS({R(fin_col)},"{lab}",{R(calc_col)},1)'
        ws.cell(r, 3).value = f'=IFERROR(B{r}/COUNTIFS({R(calc_col)},1),"—")'
        ws.cell(r, 3).number_format = "0%"
        r += 1
    ws.cell(r, 1, "Не определена")
    ws.cell(r, 2).value = (f'=COUNTIFS({R(calc_col)},1)'
                           + (f'-SUM(B{fin_first}:B{r - 1})' if r > fin_first else ""))
    ws.cell(r, 3).value = f'=IFERROR(B{r}/COUNTIFS({R(calc_col)},1),"—")'
    ws.cell(r, 3).number_format = "0%"
    r += 1
    ws.cell(r, 1, "в т.ч. определено полем Циан")
    ws.cell(r, 2).value = f'=COUNTIFS({R(src_col)},"Циан-поле",{R(calc_col)},1)'
    r += 1
    ws.cell(r, 1, "в т.ч. определено по тексту объявления")
    ws.cell(r, 2).value = f'=COUNTIFS({R(src_col)},"из описания",{R(calc_col)},1)'
    r += 2

    # — диапазоны цен —
    ws.cell(r, 1, "ДИАПАЗОН ЦЕН ПО КАТЕГОРИЯМ, ₽").font = h_font
    r += 1
    for i, h in enumerate(["Категория", "Мин. цена", "Средняя цена", "Макс. цена",
                           "Мин. ₽/м²", "Макс. ₽/м²"], start=1):
        ws.cell(r, i, h).font = h_font
    r += 1
    for cat in present_cats:
        ws.cell(r, 1, cat)
        ws.cell(r, 2).value = f'=IFERROR(_xlfn.MINIFS({R(price_col)},{R(cat_col)},"{cat}",{R(calc_col)},1),"—")'
        ws.cell(r, 3).value = f'=IFERROR(AVERAGEIFS({R(price_col)},{R(cat_col)},"{cat}",{R(calc_col)},1),"—")'
        ws.cell(r, 4).value = f'=IFERROR(_xlfn.MAXIFS({R(price_col)},{R(cat_col)},"{cat}",{R(calc_col)},1),"—")'
        ws.cell(r, 5).value = f'=IFERROR(_xlfn.MINIFS({R(ppm_col)},{R(cat_col)},"{cat}",{R(calc_col)},1),"—")'
        ws.cell(r, 6).value = f'=IFERROR(_xlfn.MAXIFS({R(ppm_col)},{R(cat_col)},"{cat}",{R(calc_col)},1),"—")'
        for cc in range(2, 7):
            ws.cell(r, cc).number_format = "#,##0"
        r += 1
    r += 1

    note = [
        "МЕТОДИКА",
        "• Срок экспозиции = сегодня − дата ПОСЛЕДНЕЙ публикации (added/creationDate). Циан сбрасывает дату при переподаче,",
        "  поэтому это «дни с последней подачи», а не полный срок размещения.",
        "• «Цена за м²» — живая формула =цена/площадь, формат «целое».",
        "• «Отделка/ремонт» берётся сначала из поля Циан (repairType/decoration), а если оно пустое — из текста объявления",
        "  по ключевым словам. Чем именно определена категория, показывает колонка «Источник отделки»; сам текст —",
        "  в колонке «Описание» (обрезан до 600 знаков, полный — по ссылке на объявление).",
        "• Полный охват (100%) выдачи Циан режется ~28 страницами × 28 лотов на запрос. Скрипт обходит лимит",
        "  дроблением по комнатности и диапазонам цены с дедупликацией по cianId; гарантированные 100% даёт только API под вашим логином.",
    ]
    for line in note:
        ws.cell(r, 1, line)
        if line == "МЕТОДИКА":
            ws.cell(r, 1).font = h_font
        r += 1
    for i, w in enumerate([26, 16, 16, 16, 14, 14], start=1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w


def _room_total_for_cat(cat, totals_by_room):
    if not totals_by_room:
        return None
    mapping = {"Студия": 9, "Своб. планировка": 7, "1": 1, "2": 2, "3": 3}
    if cat == "4+":
        return sum(v for k, v in totals_by_room.items() if k in (4, 5, 6) and v) or None
    code = mapping.get(cat)
    return totals_by_room.get(code)


# ----------------------------------------------------------------------------- #
#  PLAYWRIGHT FALLBACK (если API упёрся в капчу/блок)                            #
# ----------------------------------------------------------------------------- #

def playwright_fallback(args):
    """
    Резерв: headless Chromium открывает страницу ЖК, скроллит и парсит карточки.
    Требует `pip install playwright` и `playwright install chromium`.
    Возвращает список СЫРЫХ офферов (как API) — нормализация общая, в main().
    Из DOM достаём только cianId/ссылку/заголовок; остальные поля -> None.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.error("Playwright не установлен. `pip install playwright && playwright install chromium`")
        return []

    url = args.referer or f"https://www.cian.ru/zhiloy-kompleks-{args.jk_id}/"
    offers = []
    log.info("Playwright fallback: открываю %s", url)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=not args.headful)
        page = browser.new_page(user_agent=DEFAULT_HEADERS["User-Agent"])
        page.goto(url, wait_until="domcontentloaded", timeout=args.timeout * 1000)
        for _ in range(args.scrolls):
            page.mouse.wheel(0, 4000)
            page.wait_for_timeout(int(random.uniform(args.delay_min, args.delay_max) * 1000))
        cards = page.query_selector_all('article[data-name="CardComponent"]')
        log.info("  карточек в DOM: %d", len(cards))
        for c in cards:
            a = c.query_selector('a[href*="/sale/flat/"]')
            href = a.get_attribute("href") if a else None
            cid = None
            if href:
                m = re.search(r"/sale/flat/(\d+)/", href)
                cid = int(m.group(1)) if m else None
            title = c.inner_text().split("\n")[0] if c.inner_text() else ""
            offers.append({"cianId": cid, "fullUrl": href, "title": title})
        browser.close()
    return offers


# ----------------------------------------------------------------------------- #
#  BROWSER MODE: сбор через РЕАЛЬНЫЙ браузер (как человек)                       #
#  Запросы к API идут изнутри страницы cian.ru (fetch с credentials), поэтому    #
#  несут cookie с уже пройденным антибот-челленджем и тот же fingerprint, что у  #
#  живого пользователя. Возвращает полноценные офферы (offersSerialized).        #
# ----------------------------------------------------------------------------- #

# fetch выполняется в origin www.cian.ru -> same-origin, с cookie сессии
_JS_INPAGE_FETCH = """
async ({url, body}) => {
  try {
    const r = await fetch(url, {
      method: 'POST',
      headers: {'Content-Type': 'application/json', 'Accept': '*/*'},
      body: JSON.stringify(body),
      credentials: 'include',
    });
    const text = await r.text();
    try { return JSON.parse(text); }
    catch (e) { return {__status: r.status, __text: text.slice(0, 300)}; }
  } catch (e) { return {__error: String(e)}; }
}
"""

# мини-стелс: убрать очевидные признаки автоматизации до загрузки страницы
_JS_STEALTH = """
Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
Object.defineProperty(navigator, 'languages', {get: () => ['ru-RU', 'ru', 'en-US', 'en']});
Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
window.chrome = window.chrome || {runtime: {}};
"""


def _cookies_for_playwright(cookie_str, domain=".cian.ru"):
    return [{"name": k, "value": v, "domain": domain, "path": "/"}
            for k, v in _parse_cookie_string(cookie_str or "").items()]


class BrowserFetcher:
    """
    Тот же интерфейс, что у Fetcher (post/build_body/pause), но запрос уходит
    через page.evaluate -> in-page fetch. Благодаря этому collect_all() даёт
    room-split, price-split, дедуп и контроль охвата без изменений.
    """
    def __init__(self, args, page):
        self.args = args
        self.page = page
        self.api_url = API_URL
        self.template = None

    def build_body(self, page, room=None, price_min=None, price_max=None):
        return default_json_query(
            self.args.jk_id, self.args.region, page, self.args.engine_version,
            room, price_min, price_max, sort=getattr(self.args, "sort", None))

    def post(self, body):
        delay = self.args.backoff_base
        last = None
        for attempt in range(1, max(1, self.args.retries) + 1):
            res = self.page.evaluate(_JS_INPAGE_FETCH, {"url": self.api_url, "body": body})
            if isinstance(res, dict) and not res.get("__status") \
                    and not res.get("__error") and not res.get("__text"):
                return res
            last = res
            status = (res or {}).get("__status") or (res or {}).get("__error")
            log.warning("  браузер: ответ не JSON/блок (%s), попытка %d — пауза %.1fs",
                        status, attempt, delay)
            self.page.wait_for_timeout(int(delay * 1000))
            delay *= 2
        raise RuntimeError(f"Браузерный запрос не дал JSON: {str(last)[:200]}")

    def pause(self):
        self.page.wait_for_timeout(int(random.uniform(self.args.delay_min, self.args.delay_max) * 1000))


def browser_collect(args):
    """
    Открывает страницу ЖК в РЕАЛЬНОМ браузере (как человек), при необходимости
    подставляет ваши cookie, затем собирает все лоты через in-page API-fetch.
    Первый запуск часто требует --headful: пройти капчу руками один раз
    (потом cookie сессии валиден). Возвращает (raw_offers, totals_by_room).
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.error("Playwright не установлен: pip install playwright && playwright install chromium")
        return [], None, None

    import tempfile
    profile_dir = args.browser_profile or tempfile.mkdtemp(prefix="cian_browser_")
    jk_url = args.referer or f"https://www.cian.ru/zhiloy-kompleks-{args.jk_id}/"
    ua = args.user_agent or DEFAULT_HEADERS["User-Agent"]
    launch_kwargs = dict(
        user_data_dir=profile_dir,                        # переиспользуем профиль/сессию
        headless=not args.headful,
        user_agent=ua,
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        viewport={"width": 1920, "height": 1080},
        args=["--disable-blink-features=AutomationControlled"],
    )
    if args.browser_exec:
        launch_kwargs["executable_path"] = args.browser_exec
    if args.proxy:
        launch_kwargs["proxy"] = {"server": args.proxy}
    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(**launch_kwargs)
        ctx.add_init_script(_JS_STEALTH)
        if args.cookie:
            try:
                ctx.add_cookies(_cookies_for_playwright(args.cookie))
            except Exception as e:
                log.warning("Не удалось добавить cookie в браузер: %s", e)

        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        log.info("Браузер: открываю %s (как человек)...", jk_url)
        page.goto(jk_url, wait_until="domcontentloaded", timeout=args.timeout * 1000)

        # человеческое поведение: скролл/паузы; дать время на анти-бот/капчу
        for _ in range(3):
            page.mouse.wheel(0, int(random.uniform(1500, 4000)))
            page.wait_for_timeout(int(random.uniform(800, 2000)))
        if args.headful:
            log.info("Если показана капча — пройдите её в окне; жду %d сек...", args.captcha_wait)
            page.wait_for_timeout(args.captcha_wait * 1000)

        fetcher = BrowserFetcher(args, page)
        try:
            total_in_jk = probe_total(fetcher)
            if total_in_jk:
                log.info("Браузер: всего активных лотов в ЖК по Циан: %d", total_in_jk)
            raw, totals = collect_all(fetcher)
        finally:
            ctx.close()
    return raw, totals, total_in_jk


# ----------------------------------------------------------------------------- #
#  SELF-TEST (офлайн): прогон parse→calc→Excel на мок-данных                     #
# ----------------------------------------------------------------------------- #

MOCK_OFFERS = [
    {"cianId": 327006810, "fullUrl": "https://www.cian.ru/sale/flat/327006810/",
     "isStudio": True, "roomsCount": 0, "totalArea": "28.05", "floorNumber": 50,
     "building": {"floorsCount": 54}, "newbuilding": {"house": {"name": "Графит"}},
     "bargainTerms": {"price": 19000000}, "addedTimestamp": 1748736000,
     "editDate": "2026-06-20T10:00:00+03:00", "isByHomeowner": True,
     "user": {"userType": "homeowner"}, "decoration": "preFine",
     # поле Циан обязано победить текст: в описании евроремонт, а поле — preFine
     "description": "Сделан отличный <b>евроремонт</b>, заезжай и живи."},
    {"cianId": 327006811, "fullUrl": "https://www.cian.ru/sale/flat/327006811/",
     "isStudio": True, "roomsCount": 0, "totalArea": "27.4", "floorNumber": 31,
     "building": {"floorsCount": 54}, "newbuilding": {"house": {"name": "Сильвер"}},
     "bargainTerms": {"price": 21900000}, "addedTimestamp": 1750464000,
     "user": {"userType": "agency", "agencyName": "Илиаз и партнеры"},
     "decoration": "fine", "description": "Просторная студия с панорамным видом."},
    {"cianId": 327006812, "fullUrl": "https://www.cian.ru/sale/flat/327006812/",
     "flatType": "rooms", "roomsCount": 1, "totalArea": "41.2", "floorNumber": 12,
     "building": {"floorsCount": 25}, "newbuilding": {"house": {"name": "Кристалл"}},
     "bargainTerms": {"price": 33500000}, "addedTimestamp": 1749600000,
     "isFromBuilder": True, "user": {"userType": "developer"}, "decoration": "without",
     "description": "Передаётся без отделки. Ремонт подъезда завершён в 2025 году."},
    {"cianId": 327006813, "fullUrl": "https://www.cian.ru/sale/flat/327006813/",
     "flatType": "rooms", "roomsCount": 2, "totalArea": "60.0", "floorNumber": 8,
     "building": {"floorsCount": 36}, "newbuilding": {"house": {"name": "Сиена"}},
     "bargainTerms": {"price": 52000000}, "addedTimestamp": 1747526400,
     "editDate": "2026-06-22T09:00:00+03:00",
     "user": {"userType": "agency", "agencyName": "Элитный Дом"}, "decoration": "designer"},
    {"cianId": 327006814, "fullUrl": "https://www.cian.ru/sale/flat/327006814/",
     "flatType": "openPlan", "roomsCount": 0, "totalArea": "85.3", "floorNumber": 20,
     "building": {"floorsCount": 43}, "newbuilding": {"house": {"name": "Графит"}},
     "bargainTerms": {"price": 78000000}, "addedTimestamp": 1746230400,
     "isByHomeowner": True, "user": {"userType": "homeowner"},
     # поля отделки нет — категория берётся из текста (регресс на «авторский
     # ремонт»: раньше правило было мертво из-за \\w в JS-версии)
     "description": "Сделан авторский ремонт по дизайн-проекту, мебель остаётся."},
    {"cianId": 327006815, "fullUrl": "https://www.cian.ru/sale/flat/327006815/",
     "flatType": "rooms", "roomsCount": 4, "totalArea": "120.7", "floorNumber": 40,
     "building": {"floorsCount": 54}, "newbuilding": {"house": {"name": "Кристалл"}},
     "bargainTerms": {"price": 150000000}, "addedTimestamp": 1745020800,
     "isFromBuilder": True, "user": {"userType": "developer"}, "decoration": "fine"},
    # --- случаи, добавленные под проверку отделки из описания ---
    {"cianId": 327006816, "fullUrl": "https://www.cian.ru/sale/flat/327006816/",
     "flatType": "rooms", "roomsCount": 2, "totalArea": "54.0", "floorNumber": 3,
     "building": {"floorsCount": 12}, "bargainTerms": {"price": 24000000},
     "addedTimestamp": 1748000000,
     "description": "Квартира требуется ремонт, состояние строительное."},
    {"cianId": 327006817, "fullUrl": "https://www.cian.ru/sale/flat/327006817/",
     "flatType": "rooms", "roomsCount": 1, "totalArea": "38.0", "floorNumber": 7,
     "building": {"floorsCount": 17}, "bargainTerms": {"price": 17000000},
     "addedTimestamp": 1748100000,
     # классификатор не должен ничего выдумывать
     "description": "Продаётся квартира в новом доме. Рядом парк и школа."},
    {"cianId": 327006818, "fullUrl": "https://www.cian.ru/sale/flat/327006818/",
     "flatType": "rooms", "roomsCount": 3, "totalArea": "92.5", "floorNumber": 15,
     "building": {"floorsCount": 22}, "bargainTerms": {"price": 61000000},
     "addedTimestamp": 1748200000,
     # два реальных краша разом: \x0b роняет openpyxl (IllegalCharacterError),
     # а ведущий «=» превращает текст в формулу и Excel считает книгу битой
     "description": "=Срочно!\x0b Евроремонт. " + "очень длинный текст " * 60},
    {"cianId": 327006819, "fullUrl": "https://www.cian.ru/sale/flat/327006819/",
     "flatType": "rooms", "roomsCount": 1, "totalArea": "44.0", "floorNumber": 9,
     "building": {"floorsCount": 30}, "bargainTerms": {"price": 28000000},
     # значение поля Циан не в словаре И текста нет — отдаём значение как есть,
     # чтобы новая категория Циан не пропала молча
     "addedTimestamp": 1748300000, "decoration": "superLux"},
    {"cianId": 327006822, "fullUrl": "https://www.cian.ru/sale/flat/327006822/",
     "flatType": "rooms", "roomsCount": 2, "totalArea": "51.0", "floorNumber": 4,
     "building": {"floorsCount": 16}, "bargainTerms": {"price": 23000000},
     "addedTimestamp": 1748600000, "decoration": "superLux",
     # незнакомое значение поля УСТУПАЕТ уверенному сигналу из текста:
     # категорию из словаря мы дать не можем, а текст — можем
     "description": "Хорошее состояние, сделан свежий ремонт."},
    {"cianId": 327006820, "fullUrl": "https://www.cian.ru/sale/flat/327006820/",
     "flatType": "rooms", "roomsCount": 2, "totalArea": "58.0", "floorNumber": 2,
     "building": {"floorsCount": 9}, "bargainTerms": {"price": 21000000},
     "addedTimestamp": 1748400000, "repairType": "no",
     "description": "Квартира в жилом состоянии."},
    {"cianId": 327006821, "fullUrl": "https://www.cian.ru/sale/flat/327006821/",
     "flatType": "rooms", "roomsCount": 1, "totalArea": "40.0", "floorNumber": 5,
     "building": {"floorsCount": 14}, "bargainTerms": {"price": 19500000},
     "addedTimestamp": 1748500000,
     # стоп-контекст: ремонт относится к дому, а не к квартире
     "description": "В доме недавно сделан ремонт подъезда и заменены лифты."},
]

# Что именно должен дать классификатор: cianId -> (категория, источник).
# Ради этого словаря самотест и существует — без него мёртвое правило regex
# выглядит как успешный прогон.
EXPECT_FINISH = {
    327006810: ("Предчистовая (white box)", "Циан-поле"),   # поле важнее текста
    327006811: ("Чистовая", "Циан-поле"),
    327006812: ("Без отделки", "Циан-поле"),
    327006813: ("Дизайнерский", "Циан-поле"),               # decoration=designer
    327006814: ("Дизайнерский", "из описания"),
    327006815: ("Чистовая", "Циан-поле"),
    327006816: ("Без ремонта", "из описания"),
    327006817: (None, None),
    327006818: ("Евроремонт", "из описания"),
    327006819: ("superLux", "Циан-поле"),                   # словарь отстал от Циан
    327006822: ("Косметический", "из описания"),            # текст важнее непонятного поля
    327006820: ("Без ремонта", "Циан-поле"),                # repairType=no
    327006821: (None, None),                                # ремонт дома, не лота
}


def run_self_test(args):
    log.info("SELF-TEST: прогон parse→calc→Excel на %d мок-лотах (без сети).", len(MOCK_OFFERS))
    today = date.today()
    rows = [normalize(o, today) for o in MOCK_OFFERS]
    rows = sort_rows(rows)
    # проверка вычислений
    print("\n--- Проверка вычислений (первые строки) ---")
    print(f"{'cianId':>10} | {'кат':<6} | {'S':>6} | {'цена':>12} | {'₽/м² (calc)':>12} | "
          f"{'эксп.дн':>7} | продавец")
    for r in rows[:5]:
        ppm_check = round(r["price"] / r["area"]) if r["price"] and r["area"] else None
        ok = "OK" if ppm_check == r["ppm"] else "!!"
        print(f"{r['cianId']:>10} | {r['category']:<6} | {r['area']:>6} | "
              f"{int(r['price']):>12,} | {r['ppm']:>12,} {ok} | "
              f"{str(r['exposure_days']):>7} | {r['seller_type']}/{r['seller_name']}")
    totals = {9: 6, 7: 4, 1: 62, 2: 79, 3: 34, 4: 10}
    out = args.output or f"cian_selftest_{today.isoformat()}.xlsx"
    write_workbook(rows, args.jk_id or 2515016, args.jk_name or "SELF-TEST", totals, out)
    print(f"\nЗаписан тестовый файл: {out}")
    print_console_stats(rows)
    log_field_coverage(rows)      # заодно прогоняем диагностику заполненности

    fails = (_check_finish(rows) + _check_corpus() + _check_garbage()
             + _check_written_file(out))
    if fails:
        print(f"\nSELF-TEST ПРОВАЛЕН: расхождений {fails}")
    else:
        print("\nSELF-TEST пройден: отделка, корпус и запись файла — без расхождений.")
    return rows, fails


def _check_finish(rows):
    """Отделка и её источник на мок-лотах."""
    print("\n--- Отделка/ремонт ---")
    bad = 0
    for r in sorted(rows, key=lambda x: x["cianId"]):
        exp = EXPECT_FINISH.get(r["cianId"])
        got = (r["decoration"], r["finish_src"])
        if exp is None:
            continue
        mark = "OK" if got == exp else "!!"
        if got != exp:
            bad += 1
        print(f"  {mark} {r['cianId']} -> {got[0]!r} / {got[1]!r}"
              + ("" if got == exp else f"   ожидалось {exp[0]!r} / {exp[1]!r}"))
    return bad


def _check_garbage():
    """Мусорные значения полей из API не должны ронять прогон.

    Циан не обязан присылать строку: поле может прийти числом, словарём, списком
    или отсутствовать вовсе. Раньше нехешируемое значение падало TypeError уже
    на поиске по словарю и убивало ВЕСЬ сбор — то есть один странный оффер
    стоил пользователю всей выгрузки.
    """
    today = date.today()
    cases = [
        ("описание = None", {"cianId": 1, "description": None}),
        ("описание = число", {"cianId": 2, "description": 12345}),
        ("описание = список", {"cianId": 3, "description": ["a", "b"]}),
        ("описание = одни пробелы", {"cianId": 4, "description": "   \n\t  "}),
        ("описание = dict", {"cianId": 5, "description": {"text": "Свежий ремонт"}}),
        ("отделка = вложенный dict", {"cianId": 6, "repairType": {"value": {"name": "euro"}}}),
        ("отделка = число", {"cianId": 7, "decoration": 42}),
        ("отделка = список", {"cianId": 8, "decoration": ["fine"]}),
        ("отделка = пустая строка", {"cianId": 9, "decoration": ""}),
        ("оффер без единого поля", {}),
        ("одинокий суррогат в описании", {"cianId": 10, "description": "Ремонт \ud83d новый"}),
        ("очень длинное описание", {"cianId": 11, "description": "Ремонт. " * 9000}),
    ]
    print("\n--- Мусорные данные ---")
    bad = 0
    for name, offer in cases:
        try:
            r = normalize(offer, today)
        except Exception as e:                                  # noqa: BLE001
            print(f"  !! {name}: {type(e).__name__}: {e}")
            bad += 1
            continue
        d = r["description"]
        if d is not None and (not isinstance(d, str) or d == ""):
            print(f"  !! {name}: description = {d!r} — должно быть None или непустой строкой")
            bad += 1
            continue
        if d and len(d) > DESC_MAX_CHARS + 1:
            print(f"  !! {name}: описание длиной {len(d)} не обрезано")
            bad += 1
            continue
        if d and _SURROGATE.search(d):
            print(f"  !! {name}: в описании остался одинокий суррогат — Excel не откроет книгу")
            bad += 1
            continue
    if not bad:
        print(f"  OK все {len(cases)} случаев отработали без падения")
    return bad


def _check_corpus():
    """Общий с JS-экспортёрами корпус: tests/finish_corpus.json."""
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "tests", "finish_corpus.json")
    if not os.path.exists(path):
        print("\n--- Корпус --- пропущен: нет tests/finish_corpus.json")
        return 0
    with open(path, encoding="utf-8") as fh:
        corpus = json.load(fh)
    bad = 0
    for text, expected in corpus:
        got = finish_from_text(text)
        if got != expected:
            bad += 1
            print(f"  !! «{text}»\n     ожидалось {expected!r}, получено {got!r}")
    print(f"\n--- Корпус --- {len(corpus) - bad}/{len(corpus)}")
    if not bad:
        print("     (тот же корпус прогоняет node tests/check_finish.mjs — "
              "если оба зелёные, три экспортёра классифицируют одинаково)")
    return bad


def _check_written_file(path):
    """Обратное чтение книги: ловит то, что не видно на уровне строк —
    текст, записанный как формула, и запрещённые для Excel символы."""
    import openpyxl
    bad = 0
    wb = openpyxl.load_workbook(path)
    ws = wb["Все_лоты"]
    headers = {c.value: c.column for c in ws[4] if c.value}
    for name in ("Описание", "Отделка/ремонт", "Источник отделки"):
        if name not in headers:
            print(f"  !! в листе «Все_лоты» нет колонки «{name}»")
            bad += 1
    if "Описание" in headers:
        col = headers["Описание"]
        filled = 0
        for row in range(5, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            filled += 1
            if cell.data_type != "s":
                print(f"  !! «Описание» в строке {row} записано как {cell.data_type!r}, "
                      f"а не как текст: {str(cell.value)[:40]!r}")
                bad += 1
        print(f"\n--- Файл --- «Описание» заполнено в {filled} строках, "
              f"все ячейки текстовые: {'да' if not bad else 'НЕТ'}")
    wb.close()
    return bad


# ----------------------------------------------------------------------------- #
#  СОРТИРОВКА, СТАТИСТИКА, ИМЯ ФАЙЛА                                             #
# ----------------------------------------------------------------------------- #

def sort_rows(rows):
    """Сортировка по цене за м² по возрастанию (None — в конец)."""
    return sorted(rows, key=lambda r: (r["ppm"] is None, r["ppm"] or 0))


def print_console_stats(rows):
    ppms = [r["ppm"] for r in rows if r["ppm"]]
    exps = [r["exposure_days"] for r in rows if r["exposure_days"] is not None]
    print("\n========== ИТОГИ ==========")
    print(f"Собрано лотов: {len(rows)}")
    if ppms:
        print(f"Цена за м², ₽: мин {min(ppms):,} | средн {round(sum(ppms)/len(ppms)):,} | макс {max(ppms):,}")
    else:
        print("Цена за м²: нет данных")
    if exps:
        print(f"Срок экспозиции, дней: средний {round(sum(exps)/len(exps))} "
              f"(мин {min(exps)} / макс {max(exps)})")
    else:
        print("Срок экспозиции: нет данных")
    print("===========================")


# поля, которые НЕ всегда есть в ответе search-offers-desktop (нормально, что пустые)
_OPTIONAL_FIELDS = {"decoration", "finish_src", "description", "updated", "building",
                    "seller_name"}
# критичные поля — если пустые, скорее всего схема ответа изменилась
_CRITICAL_FIELDS = {"cianId", "url", "area", "price", "category"}


def log_field_coverage(rows):
    """
    Диагностика заполненности колонок: сразу видно, если маппинг полей сломался
    (например, Циан переименовал поле и колонка молча опустела).
    """
    n = len(rows)
    if not n:
        return
    keys = ["cianId", "url", "category", "area", "floor", "floors", "building",
            "seller_type", "seller_name", "decoration", "finish_src", "description",
            "price", "published", "updated", "exposure_days"]
    log.info("Заполненность полей (по %d лотам):", n)
    for k in keys:
        filled = sum(1 for r in rows if r.get(k) is not None)
        pct = 100 * filled / n
        flag = ""
        if filled == 0:
            flag = "  ← ПУСТО (опционально, нет в этом эндпоинте)" if k in _OPTIONAL_FIELDS \
                else "  ← ПУСТО! проверьте схему/--dump-json"
        elif k in _CRITICAL_FIELDS and pct < 90:
            flag = "  ← мало для критичного поля"
        log.info("    %-14s %3d/%-3d (%3.0f%%)%s", k, filled, n, pct, flag)
    direct = sum(1 for r in rows if r.get("url") and "/sale/flat/" in r["url"])
    log.info("    прямых ссылок /sale/flat/: %d/%d (%.0f%%)", direct, n, 100 * direct / n)

    # Разрез по источнику отделки. Без него заполненность колонки «Отделка/ремонт»
    # ничего не значит: 100% может быть следствием одного всеядного правила.
    by_field = sum(1 for r in rows if r.get("finish_src") == "Циан-поле")
    by_text = sum(1 for r in rows if r.get("finish_src") == "из описания")
    with_desc = sum(1 for r in rows if r.get("description"))
    log.info("    отделка: поле Циан %d / из описания %d / не определена %d",
             by_field, by_text, n - by_field - by_text)
    log.info("    описание есть у %d/%d (%.0f%%)", with_desc, n, 100 * with_desc / n)
    if with_desc > 0.3 * n and by_text == 0:
        log.warning("    ← описания есть, но по тексту не определено НИ ОДНОГО — "
                    "похоже, правила классификатора сломаны")
    labels = {r.get("decoration") for r in rows if r.get("finish_src") == "из описания"}
    if by_text > 0.95 * n and len(labels) == 1:
        log.warning("    ← почти всё определено по тексту и одной категорией (%s) — "
                    "похоже, в description попал общий текст ЖК, а не объявления",
                    labels.pop())


def slugify(text):
    if not text:
        return "jk"
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^0-9a-zа-яё_\-]", "", text)
    return text or "jk"


# ----------------------------------------------------------------------------- #
#  CLI                                                                           #
# ----------------------------------------------------------------------------- #

def build_argparser():
    p = argparse.ArgumentParser(
        description="Выгрузка всех активных квартир ЖК с cian.ru в Excel.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog="""\
Примеры:
  # 1) Самый надёжный путь — отдать реальный запрос из DevTools (Copy as cURL):
  #    DevTools → Network → XHR → search-offers → ПКМ → Copy as cURL → сохранить в req.txt
  python cian_scraper.py --jk "https://www.cian.ru/zhiloy-kompleks-symphony-34-2515016/" \\
                         --curl-file req.txt

  # 2) Без cURL (встроенный шаблон запроса):
  python cian_scraper.py --jk 2515016 --region 1

  # 3) Офлайн-проверка пайплайна (без сети) — пишет тестовый xlsx и печатает 5 строк:
  python cian_scraper.py --self-test
""")
    p.add_argument("--jk", help="Ссылка на ЖК или ID newobject (напр. 2515016 или URL).")
    p.add_argument("--jk-name", help="Имя ЖК для заголовков/имени файла (если не задано — берётся из данных).")
    p.add_argument("--region", type=int, default=1, help="ID региона Циан (Москва=1).")
    p.add_argument("--engine-version", type=int, default=2, help="engine_version в jsonQuery.")
    p.add_argument("--sort", default="creation_date_desc",
                   help="Стабильная сортировка выдачи (для надёжной пагинации). "
                        "Напр. creation_date_desc | price_object_order | price_object_order_desc. "
                        "При --curl-file сортировка из вашего запроса сохраняется.")
    p.add_argument("--curl-file", help="Файл с «Copy as cURL» реального запроса (headers/cookies/тело).")
    p.add_argument("--cookie", help="Строка Cookie из браузера ('k=v; k2=v2') — быстрый способ передать сессию без cURL.")
    p.add_argument("--user-agent", help="Переопределить User-Agent (полезно подставить ваш из браузера).")
    p.add_argument("--referer", help="Referer (страница ЖК). По умолчанию https://www.cian.ru/")
    p.add_argument("--output", "-o", help="Имя выходного xlsx (по умолчанию cian_<жк>_<дата>.xlsx).")

    # пагинация / обход лимита
    p.add_argument("--max-pages", type=int, default=PUBLIC_PAGE_CAP, help="Потолок страниц на один запрос.")
    p.add_argument("--no-room-split", action="store_true", help="Не дробить по комнатности (один запрос на ЖК).")
    p.add_argument("--split-price", dest="split_price", action="store_true", default=True,
                   help="Дробить по диапазонам цены при упоре в потолок (вкл).")
    p.add_argument("--no-split-price", dest="split_price", action="store_false",
                   help="Выключить дробление по цене.")
    p.add_argument("--price-min", type=int, default=None, help="Нижняя граница цены, ₽.")
    p.add_argument("--price-max", type=int, default=None, help="Верхняя граница цены, ₽.")
    p.add_argument("--price-ceiling", type=int, default=3_000_000_000,
                   help="Верхний предел для дробления, если price-max не задан.")
    p.add_argument("--min-price-span", type=int, default=500_000,
                   help="Не дробить диапазон уже этого, ₽.")

    # антибан
    p.add_argument("--delay-min", type=float, default=2.0, help="Мин. задержка между запросами, сек.")
    p.add_argument("--delay-max", type=float, default=5.0, help="Макс. задержка между запросами, сек.")
    p.add_argument("--retries", type=int, default=5, help="Число попыток с бэк-оффом на 429/403/5xx.")
    p.add_argument("--backoff-base", type=float, default=2.0, help="Базовая пауза бэк-оффа, сек (далее ×2).")
    p.add_argument("--timeout", type=float, default=30.0, help="Таймаут запроса, сек.")

    # fallback / служебное
    p.add_argument("--browser", action="store_true",
                   help="Сбор через РЕАЛЬНЫЙ браузер (как человек): API-запросы идут изнутри "
                        "страницы cian.ru с вашей сессией — проходит антибот с незаблокированного IP.")
    p.add_argument("--browser-profile", default=None,
                   help="Папка профиля браузера (хранит cookie/сессию между запусками). "
                        "По умолчанию — временный профиль.")
    p.add_argument("--browser-exec", default=None,
                   help="Путь к chromium (если не установлен через `playwright install`, "
                        "напр. в готовом окружении).")
    p.add_argument("--proxy", default=None,
                   help="Прокси для браузера, напр. http://host:port (обычно не нужен).")
    p.add_argument("--captcha-wait", type=int, default=40,
                   help="С --browser --headful: сколько секунд ждать ручного прохождения капчи.")
    p.add_argument("--playwright", action="store_true", help="Простой DOM-fallback (парсинг карточек) вместо API.")
    p.add_argument("--headful", action="store_true", help="Видимое окно браузера (нужно для прохождения капчи).")
    p.add_argument("--scrolls", type=int, default=15, help="Число прокруток страницы в Playwright-fallback.")
    p.add_argument("--dump-json", help="Сохранить сырой ответ первой страницы в файл (для сверки полей).")
    p.add_argument("--self-test", action="store_true", help="Офлайн-прогон на мок-данных (без сети).")
    p.add_argument("-v", "--verbose", action="store_true", help="Подробный лог (DEBUG).")
    return p


def main(argv=None):
    args = build_argparser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")

    today = date.today()

    if args.self_test:
        args.jk_id = extract_newobject_id(args.jk) if args.jk else 2515016
        _, fails = run_self_test(args)
        return 1 if fails else 0

    if not args.jk:
        log.error("Не задан --jk (ссылка на ЖК или ID). См. --help.")
        return 2

    args.jk_id = extract_newobject_id(args.jk)
    args.referer = args.referer or f"https://www.cian.ru/zhiloy-kompleks-{args.jk_id}/"
    log.info("ЖК ID (newobject) = %s, регион = %s", args.jk_id, args.region)

    # Источник данных -> всегда СЫРЫЕ офферы (raw); нормализация единым проходом ниже.
    total_in_jk = None
    totals_by_room = None
    if args.browser:
        try:
            raw, totals_by_room, total_in_jk = browser_collect(args)
        except RuntimeError as e:
            log.error("Браузерный сбор не удался (%s). Откройте с --headful и пройдите капчу, "
                      "или передайте --cookie/--curl-file.", e)
            raw = []
    elif args.playwright:
        raw = playwright_fallback(args)
    else:
        curl = None
        if args.curl_file:
            with open(args.curl_file, "r", encoding="utf-8") as f:
                curl = parse_curl(f.read())
            log.info("Загружен cURL: %d заголовков, %d cookie, тело %s",
                     len(curl["headers"]), len(curl["cookies"]),
                     "есть" if curl["json"] else "нет (будет шаблон)")
        fetcher = Fetcher(args, curl)

        # опциональный дамп сырого ответа первой страницы
        if args.dump_json:
            try:
                resp = fetcher.post(fetcher.build_body(1))
                with open(args.dump_json, "w", encoding="utf-8") as f:
                    json.dump(resp, f, ensure_ascii=False, indent=2)
                log.info("Сырой ответ первой страницы сохранён в %s", args.dump_json)
            except Exception as e:
                log.error("Не удалось получить/сохранить ответ: %s", e)

        # сколько всего лотов в ЖК (для контроля охвата)
        total_in_jk = probe_total(fetcher)
        if total_in_jk:
            log.info("Всего активных лотов в ЖК по Циан: %d", total_in_jk)

        try:
            raw, totals_by_room = collect_all(fetcher)
        except RuntimeError as e:
            log.error("API недоступен (%s). Пробую Playwright-fallback...", e)
            raw = playwright_fallback(args)

    rows = [normalize(o, today) for o in raw]

    if not rows:
        log.error("Не собрано ни одного лота. Проверьте --jk, cookie/--curl-file "
                  "или используйте --playwright.")
        return 1

    # имя ЖК из данных, если не задано (для Playwright-офферов полей нет -> None)
    jk_name = args.jk_name
    if not jk_name and raw:
        jk_name = dig(raw[0], "newbuilding.name") or dig(raw[0], "geo.jk.name")
    jk_name = jk_name or f"JK {args.jk_id}"

    rows = sort_rows(rows)

    # диагностика заполненности колонок (видно сразу, если маппинг сломан)
    log_field_coverage(rows)

    # контроль охвата: собрали ли мы все лоты ЖК
    direct = sum(1 for r in rows if r.get("url") and "/sale/flat/" in r["url"])
    if total_in_jk:
        cov = 100.0 * len(rows) / total_in_jk
        log.info("Охват: собрано %d из %d (%.0f%%); прямых ссылок на лот: %d/%d",
                 len(rows), total_in_jk, cov, direct, len(rows))
        if len(rows) < total_in_jk:
            log.warning("Собрано меньше, чем всего в ЖК. Попробуйте --split-price "
                        "и/или увеличьте --max-pages; проверьте свежесть cookie.")
    else:
        log.info("Собрано %d лотов; прямых ссылок на лот: %d/%d", len(rows), direct, len(rows))

    out = args.output or f"cian_{slugify(jk_name)}_{today.isoformat()}.xlsx"
    write_workbook(rows, args.jk_id, jk_name, totals_by_room, out, total_in_jk=total_in_jk)
    log.info("Записан файл: %s", out)

    # консоль (значения могут быть None — в Playwright-fallback цена/площадь отсутствуют)
    print("\n--- Первые 5 строк (после сортировки по ₽/м²) ---")
    for r in rows[:5]:
        price = f"{int(r['price']):,}" if r['price'] is not None else "—"
        ppm = f"{r['ppm']:,}" if r['ppm'] is not None else "—"
        area = r['area'] if r['area'] is not None else "—"
        exp = r['exposure_days'] if r['exposure_days'] is not None else "—"
        print(f"  {r['cianId']} | {r['category']} | {area} м² | "
              f"{price} ₽ | {ppm} ₽/м² | эксп. {exp} дн | {r['url']}")
    print_console_stats(rows)
    return 0


if __name__ == "__main__":
    sys.exit(main())
